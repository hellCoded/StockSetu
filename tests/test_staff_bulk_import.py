import io
import pytest
import openpyxl
from werkzeug.datastructures import FileStorage
from inventory_app.services.auth_service import import_staff_bulk, generate_staff_template, authenticate_user
from inventory_app.database import get_db


def test_generate_staff_template_xlsx():
    mem, filename, mimetype = generate_staff_template(file_format="xlsx")
    assert filename.endswith(".xlsx")
    assert mimetype == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    wb = openpyxl.load_workbook(mem)
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    assert rows[0] == ("Employee ID", "Full Name", "Phone No", "Email")
    assert len(rows) >= 5


def test_generate_staff_template_csv():
    mem, filename, mimetype = generate_staff_template(file_format="csv")
    assert filename.endswith(".csv")
    assert mimetype == "text/csv"
    content = mem.getvalue().decode("utf-8")
    assert "Employee ID,Full Name,Phone No,Email" in content


def test_import_staff_bulk_csv_success(app):
    with app.app_context():
        csv_data = (
            "Employee ID,Full Name,Phone No,Email\n"
            "EMP-2001,Alice Smith,9876543210,alice@company.com\n"
            "EMP-2002,Bob Jones,9876543211,bob@company.com\n"
        )
        file_storage = FileStorage(
            stream=io.BytesIO(csv_data.encode("utf-8")),
            filename="staff_import.csv",
            content_type="text/csv"
        )
        success, msg, details = import_staff_bulk(file_storage, default_password="CustomPassword123", imported_by="testadmin")
        assert success is True
        assert details["imported_count"] == 2
        assert details["role_request_count"] == 0
        
        db = get_db()
        user_alice = db.users.find_one({"employee_id": "EMP-2001"})
        assert user_alice is not None
        assert user_alice["name"] == "Alice Smith"
        assert user_alice["phone"] == "9876543210"
        assert user_alice["email"] == "alice@company.com"
        assert user_alice["role"] == "staff"
        assert user_alice["is_active"] is False
        
        # Test authenticating with Employee ID
        auth_success, auth_msg, auth_user = authenticate_user("EMP-2001", "CustomPassword123")
        assert auth_success is True
        assert auth_user["employee_id"] == "EMP-2001"


def test_import_staff_bulk_xlsx_success(app):
    with app.app_context():
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Employee ID", "Full Name", "Phone No", "Email"])
        ws.append(["EMP-3001", "Charlie Brown", "9812345678", "charlie@company.com"])
        
        excel_bytes = io.BytesIO()
        wb.save(excel_bytes)
        excel_bytes.seek(0)
        
        file_storage = FileStorage(
            stream=excel_bytes,
            filename="staff.xlsx",
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        success, msg, details = import_staff_bulk(file_storage, default_password="DefaultPass123", imported_by="testadmin")
        assert success is True
        assert details["imported_count"] == 1
        
        db = get_db()
        user = db.users.find_one({"employee_id": "EMP-3001"})
        assert user is not None
        assert user["name"] == "Charlie Brown"
        assert user["phone"] == "9812345678"
        assert user["role"] == "staff"


def test_import_staff_bulk_skip_duplicate_and_invalid(app):
    with app.app_context():
        csv_data = (
            "Employee ID,Full Name,Phone No,Email\n"
            "EMP-9999,Admin User,9999999999,admin@test.com\n" # Existing email in seeded db
            "EMP-8888,,8888888888,\n"                         # Missing name/identifier
            "EMP-7777,David Valid,7777777777,david@company.com\n" # Valid
        )
        file_storage = FileStorage(
            stream=io.BytesIO(csv_data.encode("utf-8")),
            filename="mixed.csv",
            content_type="text/csv"
        )
        success, msg, details = import_staff_bulk(file_storage, imported_by="testadmin")
        assert success is True
        assert details["imported_count"] == 1
        assert details["skipped_count"] == 2


def test_bulk_import_route_admin(admin_client):
    csv_data = (
        "Employee ID,Full Name,Phone No,Email\n"
        "EMP-5001,Elena Rostova,9555123456,elena@company.com\n"
    )
    data = {
        "staff_file": (io.BytesIO(csv_data.encode("utf-8")), "staff.csv"),
        "default_password": "Staff@123"
    }
    resp = admin_client.post("/users/bulk-import", data=data, content_type="multipart/form-data", follow_redirects=True)
    assert resp.status_code == 200
    assert b"Successfully imported 1 staff member" in resp.data


def test_download_template_route(admin_client):
    resp = admin_client.get("/users/template?format=xlsx")
    assert resp.status_code == 200
    assert resp.headers["Content-Disposition"].startswith("attachment;")
    assert "StockSetu_Staff_Import_Template.xlsx" in resp.headers["Content-Disposition"]


def test_bulk_import_route_unauthorized(staff_client):
    resp = staff_client.post("/users/bulk-import", data={}, follow_redirects=True)
    assert b"Forbidden" in resp.data or resp.status_code == 403


