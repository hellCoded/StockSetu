"""
Export Service Module
---------------------
Centralized utility module for generating standardized Excel (.xlsx) 
and PDF (.pdf) reports across the StockSetu platform.

Provides reusable components for:
- OpenPyXL workbook formatting (headers, alternating row colors, summary rows, freeze panes, auto-filters)
- ReportLab PDF document layout (A4 landscape, KPI blocks, custom styled tables, page numbering)
"""

import io
from datetime import datetime, timezone
from typing import List, Dict, Any, Optional, Tuple

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfgen import canvas


def get_product_export_filters(request) -> dict:
    """Extract and normalize product export filter parameters from request."""
    query = request.args.get('q', '').strip()
    category = request.args.get('category', '').strip()
    location = request.args.get('location', '').strip()
    stock_status = request.args.get('status', '').strip()
    show_inactive = request.args.get('show_inactive', '0') == '1'
    
    is_active_filter = None if show_inactive else True
    
    return {
        'query': query,
        'category': category,
        'location': location,
        'stock_status': stock_status,
        'is_active': is_active_filter,
        'show_inactive': show_inactive
    }


def get_transaction_export_filters(request) -> dict:
    """Extract and normalize transaction export filter parameters from request."""
    product_name = request.args.get('product_name', '').strip()
    tx_type = request.args.get('type', '').strip()
    
    return {
        'product_name': product_name,
        'tx_type': tx_type
    }


def build_export_subtitle(base_text: str, filters: dict, record_count: int) -> str:
    """Build standardized export subtitle with timestamp and active filters."""
    now_str = datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S UTC')
    return f"Generated on {now_str} | {base_text} | Total Records: {record_count}"


class NumberedCanvas(canvas.Canvas):
    """
    Two-pass canvas to dynamically compute and render total page count
    and footer confidentiality notices on every page of a ReportLab PDF.
    """
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._saved_page_states: List[Dict[str, Any]] = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.draw_page_footer(num_pages)
            super().showPage()
        super().save()

    def draw_page_footer(self, page_count: int):
        self.saveState()
        self.setFont("Helvetica", 8)
        self.setFillColor(colors.HexColor("#64748B"))
        self.setStrokeColor(colors.HexColor("#CBD5E1"))
        self.setLineWidth(0.5)
        self.line(36, 28, 805, 28)
        
        self.drawString(36, 16, "Confidential — StockSetu Executive Report")
        page_text = f"Page {self._pageNumber} of {page_count}"
        self.drawRightString(805, 16, page_text)
        self.restoreState()


def generate_excel_export(
    sheet_title: str,
    banner_title: str,
    subtitle: str,
    headers: List[str],
    data_rows: List[List[Any]],
    col_widths: Dict[int, int],
    number_formats: Optional[Dict[int, str]] = None,
    alignments: Optional[Dict[int, str]] = None,
    status_col_idx: Optional[int] = None,
    status_style_map: Optional[Dict[str, Tuple[PatternFill, Font]]] = None,
    summary_config: Optional[Dict[str, Any]] = None
) -> io.BytesIO:
    """
    Builds a standardized Excel spreadsheet in A4 landscape layout.

    :param sheet_title: Title of the active worksheet
    :param banner_title: Main header title banner string
    :param subtitle: Subtitle string containing date and active filters
    :param headers: Column header labels
    :param data_rows: List of row values
    :param col_widths: Dict mapping column index (1-based) to width in characters
    :param number_formats: Optional dict mapping column index to openpyxl number_format
    :param alignments: Optional dict mapping column index to alignment ('left', 'center', 'right')
    :param status_col_idx: Optional 1-based column index for status badges
    :param status_style_map: Optional dict mapping status string to (fill, font) tuple
    :param summary_config: Optional dict with 'label_col', 'sum_cols' (list of col idxs)
    :return: BytesIO buffer containing the Excel binary file
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title
    ws.views.sheetView[0].showGridLines = True

    # Page Setup for A4 Landscape
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    font_family = "Segoe UI"
    title_font = Font(name=font_family, size=15, bold=True, color="FFFFFF")
    title_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    sub_font = Font(name=font_family, size=9, italic=True, color="94A3B8")
    header_font = Font(name=font_family, size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    bold_font = Font(name=font_family, size=10, bold=True, color="0F172A")
    regular_font = Font(name=font_family, size=10, color="1E293B")
    
    fill_even = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    fill_odd = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )
    total_border = Border(
        top=Side(style='thin', color='0F172A'),
        bottom=Side(style='double', color='0F172A'),
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0')
    )

    num_cols = len(headers)
    last_col_letter = get_column_letter(num_cols)

    # 1. Header Title Banner (Row 1 & 2)
    ws.merge_cells(f'A1:{last_col_letter}1')
    t_cell = ws['A1']
    t_cell.value = banner_title
    t_cell.font = title_font
    t_cell.fill = title_fill
    t_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells(f'A2:{last_col_letter}2')
    s_cell = ws['A2']
    s_cell.value = subtitle
    s_cell.font = sub_font
    s_cell.fill = title_fill
    s_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    # 2. Table Header Row (Row 4)
    header_row = 4
    ws.row_dimensions[header_row].height = 26
    for col_num, h_text in enumerate(headers, 1):
        c = ws.cell(row=header_row, column=col_num, value=h_text)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = thin_border

    # 3. Data Rows (Row 5+)
    start_row = 5
    for idx, row_vals in enumerate(data_rows, start=start_row):
        ws.append(row_vals)
        r_fill = fill_even if idx % 2 == 0 else fill_odd
        ws.row_dimensions[idx].height = 20
        
        for col_num in range(1, num_cols + 1):
            cell = ws.cell(row=idx, column=col_num)
            cell.font = regular_font
            cell.fill = r_fill
            cell.border = thin_border

            # Alignments
            if alignments and col_num in alignments:
                cell.alignment = Alignment(horizontal=alignments[col_num], vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

            # Number Formats
            if number_formats and col_num in number_formats:
                cell.number_format = number_formats[col_num]

            # Status Badge Styling
            if status_col_idx and col_num == status_col_idx and status_style_map:
                st_val = str(cell.value or '').strip()
                if st_val in status_style_map:
                    st_fill, st_font = status_style_map[st_val]
                    cell.fill = st_fill
                    cell.font = st_font

    end_data_row = start_row + len(data_rows) - 1

    # 4. Summary Total Row (if configured and data exists)
    if len(data_rows) > 0 and summary_config:
        tot_row_idx = end_data_row + 1
        ws.row_dimensions[tot_row_idx].height = 24

        label_col = summary_config.get('label_col', 1)
        tot_label = ws.cell(row=tot_row_idx, column=label_col, value="TOTAL SUMMARY")
        tot_label.font = bold_font
        tot_label.alignment = Alignment(horizontal="left", vertical="center")

        sum_cols = summary_config.get('sum_cols', [])
        for sc in sum_cols:
            col_let = get_column_letter(sc)
            sc_cell = ws.cell(row=tot_row_idx, column=sc, value=f"=SUM({col_let}{start_row}:{col_let}{end_data_row})")
            sc_cell.font = bold_font
            sc_cell.alignment = Alignment(horizontal="right", vertical="center")
            if number_formats and sc in number_formats:
                sc_cell.number_format = number_formats[sc]

        for c in range(1, num_cols + 1):
            cell = ws.cell(row=tot_row_idx, column=c)
            cell.border = total_border

        ws.auto_filter.ref = f"A4:{last_col_letter}{end_data_row}"

    ws.freeze_panes = "A5"

    # 5. Set Custom Column Widths
    for c_idx, width in col_widths.items():
        c_let = get_column_letter(c_idx)
        ws.column_dimensions[c_let].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generate_pdf_export(
    title: str,
    subtitle: str,
    headers: List[str],
    data_rows: List[List[Any]],
    col_widths: List[float],
    kpi_summary: Optional[List[List[str]]] = None
) -> io.BytesIO:
    """
    Builds a standardized ReportLab PDF document in A4 landscape format.

    :param title: PDF Title Banner String
    :param subtitle: PDF Subtitle Filter Info String
    :param headers: Table column header labels
    :param data_rows: List of formatted table row Paragraph objects/strings
    :param col_widths: List of column widths in points (total ~770pt for A4 landscape)
    :param kpi_summary: Optional list of 2 rows for KPI summary block
    :return: BytesIO buffer containing PDF binary file
    """
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=36,
        leftMargin=36,
        topMargin=36,
        bottomMargin=42
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=16,
        textColor=colors.HexColor('#0F172A'),
        spaceAfter=3
    )
    subtitle_style = ParagraphStyle(
        'DocSubTitle',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9,
        textColor=colors.HexColor('#64748B'),
        spaceAfter=10
    )
    table_header_style = ParagraphStyle(
        'TableHeader',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=8.5,
        textColor=colors.white,
        alignment=1
    )
    cell_center = ParagraphStyle(
        'CellCenter',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=8.5,
        alignment=1
    )

    elements = []

    # Title & Subtitle
    elements.append(Paragraph(title, title_style))
    elements.append(Paragraph(subtitle, subtitle_style))

    # Optional KPI Summary Block
    if kpi_summary:
        summary_paragraphs = [
            [Paragraph(f"<b>{cell}</b>", cell_center) for cell in row]
            for row in kpi_summary
        ]
        sum_col_w = [770.0 / len(kpi_summary[0])] * len(kpi_summary[0])
        summary_table = Table(summary_paragraphs, colWidths=sum_col_w)
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#F1F5F9')),
            ('BACKGROUND', (0,1), (-1,1), colors.HexColor('#FFFFFF')),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E1')),
            ('TOPPADDING', (0,0), (-1,-1), 4),
            ('BOTTOMPADDING', (0,0), (-1,-1), 4),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE')
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 10))

    # Data Table Setup
    header_row_paras = [Paragraph(h, table_header_style) for h in headers]
    table_data = [header_row_paras] + data_rows

    doc_table = Table(table_data, colWidths=col_widths, repeatRows=1)
    
    t_style = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E293B')),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
    ]

    for r_idx in range(1, len(table_data)):
        bg_col = colors.HexColor('#F8FAFC') if r_idx % 2 == 0 else colors.HexColor('#FFFFFF')
        t_style.append(('BACKGROUND', (0, r_idx), (-1, r_idx), bg_col))

    doc_table.setStyle(TableStyle(t_style))
    elements.append(doc_table)

    doc.build(elements, canvasmaker=NumberedCanvas)
    buffer.seek(0)
    return buffer
