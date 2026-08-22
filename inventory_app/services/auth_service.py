import io
import csv
import json
import logging
from datetime import datetime, timezone
from bson import ObjectId
from werkzeug.security import generate_password_hash, check_password_hash
from inventory_app.database import get_db
from inventory_app import cache_get, cache_set, cache_delete

logger = logging.getLogger(__name__)

def invalidate_user_cache(user_id: str):
    """Invalidates cached user session data."""
    if user_id:
        cache_delete(f"auth:user:{str(user_id)}")

def register_user(email: str = "", password: str = "", role: str = "staff", name: str = "", surname: str = "", employee_id: str = "", phone: str = "") -> tuple[bool, str, dict]:
    """Registers a new user with employee_id as the primary unique business key."""
    import re
    from inventory_app.utils.validators import generate_employee_id
    db = get_db()
    
    clean_emp_id = (employee_id or "").strip()
    if not clean_emp_id:
        clean_emp_id = generate_employee_id(name)
        
    clean_email = email.strip()
    clean_name = name.strip()
    clean_surname = surname.strip() if surname else ""
    clean_phone = phone.strip() if phone else ""
    full_name = f"{clean_name} {clean_surname}".strip() if clean_surname else clean_name
    
    # Check duplicate employee_id via indexed lowercase field
    if db.users.find_one({"employee_id_lower": clean_emp_id}):
        return False, f"Employee ID '{clean_emp_id}' is already registered.", {}
        
    # Check duplicate email
    if db.users.find_one({"email": clean_email}):
        return False, "Email address is already registered.", {}
        
    valid_roles = ["admin", "inventory_manager", "staff"]
    assigned_role = role if role in valid_roles else "staff"
    
    now = datetime.now(timezone.utc)
    user_doc = {
        "employee_id": clean_emp_id,
        "employee_id_lower": clean_emp_id,
        "name": full_name or clean_emp_id,
        "name_lower": full_name or clean_emp_id,
        "email": clean_email,
        "email_lower": clean_email,
        "phone": clean_phone,
        "password_hash": generate_password_hash(password),
        "role": assigned_role,
        "is_active": True,
        "last_active_at": now,
        "created_at": now,
        "updated_at": now
    }
    
    try:
        result = db.users.insert_one(user_doc)
        user_doc["_id"] = str(result.inserted_id)
        return True, "Registration successful.", user_doc
    except Exception as e:
        return False, f"Failed to create user account: {str(e)}", {}

def set_user_active_status(user_id: str, is_active: bool):
    """Sets user is_active boolean state in database."""
    db = get_db()
    try:
        update_op = {"$set": {"is_active": is_active, "updated_at": datetime.now(timezone.utc)}}
        if not is_active:
            update_op["$unset"] = {"session_token": ""}
        db.users.update_one(
            {"_id": ObjectId(user_id)},
            update_op
        )
        invalidate_user_cache(user_id)
        from inventory_app import cache_delete
        cache_delete("dashboard:main")
    except Exception as e:
        logger.error(f"Failed to set user active status: {e}")

def record_user_activity(user_id: str):
    """Updates the last_active_at timestamp for a user and marks them active."""
    uid_str = str(user_id) if user_id else ""
    if not uid_str or not ObjectId.is_valid(uid_str):
        return
    db = get_db()
    try:
        now = datetime.now(timezone.utc)
        db.users.update_one(
            {"_id": ObjectId(uid_str)},
            {"$set": {"last_active_at": now, "updated_at": now}}
        )
    except Exception as e:
        logger.error(f"Failed to record user activity: {e}")

def deactivate_inactive_users(inactivity_hours: float = 12.0) -> int:
    """
    Finds all active users who have been inactive/offline for more than `inactivity_hours`
    and marks them as inactive (is_active = False).
    Returns count of users deactivated.
    """
    from datetime import timedelta
    db = get_db()
    cutoff = datetime.now(timezone.utc) - timedelta(hours=inactivity_hours)
    try:
        result = db.users.update_many(
            {
                "is_active": True,
                "$or": [
                    {"last_active_at": {"$lt": cutoff}},
                    {"last_active_at": {"$exists": False}, "updated_at": {"$lt": cutoff}},
                    {"last_active_at": {"$exists": False}, "updated_at": {"$exists": False}, "created_at": {"$lt": cutoff}}
                ]
            },
            {"$set": {"is_active": False, "updated_at": datetime.now(timezone.utc)}}
        )
        if result.modified_count > 0:
            from inventory_app import cache_delete
            cache_delete("dashboard:main")
        return result.modified_count
    except Exception as e:
        logger.error(f"Failed to deactivate inactive users: {e}")
        return 0

def authenticate_user(identifier: str, password: str) -> tuple[bool, str, dict]:
    """Authenticates a user via employee_id or email and returns the user document."""
    import re
    db = get_db()
    clean_id = (identifier or "").strip()
    if not clean_id or not password:
        return False, "Employee ID/Email and password are required.", {}
        
    user = db.users.find_one({
        "$or": [
            {"employee_id_lower": clean_id.lower()},
            {"employee_id": clean_id},
            {"employee_id": {"$regex": f"^{re.escape(clean_id)}$", "$options": "i"}},
            {"email_lower": clean_id.lower()},
            {"email": clean_id.lower()}
        ]
    })
    
    if not user:
        return False, "Account not found. Please register an account first.", {}
        
    if check_password_hash(user.get("password_hash", ""), password):
        now = datetime.now(timezone.utc)
        db.users.update_one(
            {"_id": user["_id"]},
            {"$set": {"is_active": True, "last_active_at": now, "updated_at": now}}
        )
        user["is_active"] = True
        user["last_active_at"] = now
        user["_id"] = str(user["_id"])
        return True, "Login successful.", user
        
    return False, "Invalid Employee ID/Email or password.", {}

def get_all_users(search: str = "", role: str = "", page: int = None, per_page: int = 25, return_total: bool = False):
    """Retrieves users for admin management with optional filtering and pagination.
    Excludes sensitive authentication fields (password_hash, session_token) from results.
    """
    db = get_db()
    query = {}
    if search:
        import re
        escaped = re.escape(search.strip())
        # Use indexed lowercase fields for faster search where available,
        # fall back to $regex on non-indexed fields (phone, surname)
        query["$or"] = [
            {"employee_id_lower": {"$regex": escaped, "$options": "i"}},
            {"email_lower": {"$regex": escaped, "$options": "i"}},
            {"name_lower": {"$regex": escaped, "$options": "i"}},
            {"phone": {"$regex": escaped, "$options": "i"}},
        ]
    if role:
        query["role"] = role.strip().lower()

    effective_page = max(1, int(page or 1)) if page is not None else 1
    effective_limit = max(1, int(per_page or 25)) if per_page is not None else 25
    effective_skip = (effective_page - 1) * effective_limit

    # Single-query pagination: fetch limit+1 to detect "has more"
    fetch_limit = effective_limit + 1 if return_total else effective_limit

    # Projection: exclude sensitive auth fields from user-list queries
    projection = {
        "password_hash": 0,
        "session_token": 0,
    }

    cursor = db.users.find(query, projection).sort("created_at", -1)
    cursor = cursor.skip(effective_skip).limit(fetch_limit)

    users = list(cursor)
    has_more = len(users) > effective_limit
    if has_more:
        users = users[:effective_limit]
    for u in users:
        u["_id"] = str(u["_id"])

    if return_total:
        # When has_more is False we fetched ≤ effective_limit from a limit+1 query,
        # so we are on the last page and the true total is skip + items returned.
        total_count = (effective_skip + len(users)) if not has_more else db.users.count_documents(query)
        return users, total_count
    return users


def get_user_stats() -> dict:
    """Computes total counts of active, inactive, admin, manager, and staff accounts."""
    db = get_db()
    pipeline = [
        {"$group": {
            "_id": None,
            "total": {"$sum": 1},
            "active": {"$sum": {"$cond": [{"$eq": ["$is_active", True]}, 1, 0]}},
            "inactive": {"$sum": {"$cond": [{"$ne": ["$is_active", True]}, 1, 0]}},
            "admins": {"$sum": {"$cond": [{"$eq": ["$role", "admin"]}, 1, 0]}},
            "managers": {"$sum": {"$cond": [{"$eq": ["$role", "inventory_manager"]}, 1, 0]}},
            "staff": {"$sum": {"$cond": [{"$eq": ["$role", "staff"]}, 1, 0]}},
        }}
    ]
    res = list(db.users.aggregate(pipeline))
    if res:
        stats = res[0]
        stats.pop("_id", None)
        return stats
    return {"total": 0, "active": 0, "inactive": 0, "admins": 0, "managers": 0, "staff": 0}


def get_user_by_id(user_id: str):
    """Retrieves user by string ID (cached 30s for session validation)."""
    uid_str = str(user_id) if user_id else ""
    if not uid_str or not ObjectId.is_valid(uid_str):
        return None
    cache_key = f"auth:user:{uid_str}"
    cached = cache_get(cache_key)
    if cached is not None:
        try:
            return json.loads(cached)
        except Exception:
            pass
    db = get_db()
    try:
        user = db.users.find_one({"_id": ObjectId(uid_str)})
        if user:
            user["_id"] = str(user["_id"])
            cache_set(cache_key, json.dumps(user, default=str), ttl=30)
        return user
    except Exception:
        return None

def update_user_role(user_id: str, new_role: str) -> tuple[bool, str]:
    """Updates user role (Admin only)."""
    uid_str = str(user_id) if user_id else ""
    if not uid_str or not ObjectId.is_valid(uid_str):
        return False, "Invalid user ID."
    db = get_db()
    valid_roles = ["admin", "inventory_manager", "staff"]
    if new_role not in valid_roles:
        return False, "Invalid role specified."
        
    try:
        result = db.users.update_one(
            {"_id": ObjectId(uid_str)},
            {"$set": {"role": new_role, "updated_at": datetime.now(timezone.utc)}}
        )
        if result.modified_count > 0 or result.matched_count > 0:
            invalidate_user_cache(uid_str)
            from inventory_app import cache_delete
            cache_delete("dashboard:main")
            return True, f"User role updated to {new_role}."
        return False, "User not found."
    except Exception as e:
        return False, f"Failed to update role: {str(e)}"

def toggle_user_active(user_id: str) -> tuple[bool, str, bool]:
    """Toggles user active state (Admin only)."""
    uid_str = str(user_id) if user_id else ""
    if not uid_str or not ObjectId.is_valid(uid_str):
        return False, "Invalid user ID.", False
    db = get_db()
    try:
        user = db.users.find_one({"_id": ObjectId(uid_str)})
        if not user:
            return False, "User not found.", False
            
        new_status = not user.get("is_active", True)
        db.users.update_one(
            {"_id": ObjectId(uid_str)},
            {"$set": {"is_active": new_status, "updated_at": datetime.now(timezone.utc)}}
        )
        invalidate_user_cache(uid_str)
        from inventory_app import cache_delete
        cache_delete("dashboard:main")
        status_str = "activated" if new_status else "deactivated"
        return True, f"User account has been {status_str}.", new_status
    except Exception as e:
        return False, f"Failed to update user status: {str(e)}", False

def change_password(user_id: str, old_password: str, new_password: str) -> tuple[bool, str]:
    """Changes password for logged in user."""
    uid_str = str(user_id) if user_id else ""
    if not uid_str or not ObjectId.is_valid(uid_str):
        return False, "Invalid user ID."
    db = get_db()
    try:
        user = db.users.find_one({"_id": ObjectId(uid_str)})
        if not user:
            return False, "User not found."
            
        if not check_password_hash(user.get("password_hash", ""), old_password):
            return False, "Current password is incorrect."
            
        if len(new_password) < 6:
            return False, "New password must be at least 6 characters long."
            
        db.users.update_one(
            {"_id": ObjectId(uid_str)},
            {"$set": {
                "password_hash": generate_password_hash(new_password),
                "updated_at": datetime.now(timezone.utc)
            },
            "$unset": {"session_token": ""}}
        )
        invalidate_user_cache(uid_str)
        return True, "Password changed successfully. Please log in again on all devices."
    except Exception as e:
        return False, f"Failed to change password: {str(e)}"

def update_user_profile_info(user_id: str, name: str, email: str) -> tuple[bool, str]:
    """Updates name and email for logged-in user profile."""
    uid_str = str(user_id) if user_id else ""
    if not uid_str or not ObjectId.is_valid(uid_str):
        return False, "Invalid user ID."
    db = get_db()
    if not email or '@' not in email:
        return False, "Please provide a valid email address."
        
    try:
        existing = db.users.find_one({"email": email.strip(), "_id": {"$ne": ObjectId(uid_str)}})
        if existing:
            return False, "This email address is already registered to another account."
            
        db.users.update_one(
            {"_id": ObjectId(uid_str)},
            {"$set": {
                "name": name.strip(),
                "email": email.strip(),
                "updated_at": datetime.now(timezone.utc)
            }}
        )
        invalidate_user_cache(uid_str)
        return True, "Profile details updated successfully."
    except Exception as e:
        return False, f"Failed to update profile: {str(e)}"

def re_escape(text: str) -> str:
    import re
    return re.escape(text)

def create_role_request(user_id: str, employee_id: str = "", email: str = "", requested_role: str = "inventory_manager", reason: str = "") -> tuple[bool, str]:
    """Submits a role promotion/update request from user to admin for verification."""
    uid_str = str(user_id) if user_id else ""
    if not uid_str or not ObjectId.is_valid(uid_str):
        return False, "Invalid user ID."
    db = get_db()
    valid_requested_roles = ["inventory_manager", "admin"]
    if requested_role not in valid_requested_roles:
        return False, "Invalid requested role."

    if not reason or not reason.strip():
        return False, "A reason / justification is required to submit a role request."

    effective_emp_id = (employee_id or "").strip()

    try:
        existing = db.role_requests.find_one({"user_id": uid_str, "status": "PENDING"})
        if existing:
            return False, "You already have a pending request"
            
        user = db.users.find_one({"_id": ObjectId(uid_str)})
        current_role = user.get("role", "staff") if user else "staff"
        if not effective_emp_id and user:
            effective_emp_id = user.get("employee_id", "")

        if current_role == requested_role:
            return False, f"You are already assigned the '{current_role}' role."

        now = datetime.now(timezone.utc)
        request_doc = {
            "user_id": uid_str,
            "employee_id": effective_emp_id,
            "email": email or (user.get("email") if user else ""),
            "current_role": current_role,
            "requested_role": requested_role,
            "reason": reason.strip(),
            "status": "PENDING",
            "created_at": now,
            "updated_at": now
        }
        db.role_requests.insert_one(request_doc)
        
        from inventory_app.services.audit_service import log_audit
        log_audit("role_request_submitted", effective_emp_id, target_resource=uid_str, details={"requested_role": requested_role, "reason": reason.strip()})
        
        from inventory_app import cache_delete
        cache_delete("dashboard:main")
        
        return True, f"Role request ({requested_role.replace('_', ' ').title()}) submitted to Administrator for verification."
    except Exception as e:
        return False, f"Failed to submit role request: {str(e)}"

def cancel_role_request(request_id: str, user_id: str) -> tuple[bool, str]:
    """Withdraws/cancels a pending role request."""
    req_id_str = str(request_id) if request_id else ""
    uid_str = str(user_id) if user_id else ""
    if not req_id_str or not ObjectId.is_valid(req_id_str):
        return False, "Invalid request ID."
    db = get_db()
    try:
        req = db.role_requests.find_one({"_id": ObjectId(req_id_str)})
        if not req:
            return False, "Role request not found."
            
        if str(req.get("user_id")) != uid_str:
            return False, "You are not authorized to cancel this request."
            
        if req.get("status") != "PENDING":
            return False, "Only pending requests can be cancelled."
            
        now = datetime.now(timezone.utc)
        db.role_requests.update_one(
            {"_id": ObjectId(req_id_str)},
            {"$set": {"status": "CANCELLED", "updated_at": now}}
        )
        
        from inventory_app.services.audit_service import log_audit
        log_audit("role_request_cancelled", req.get("employee_id"), target_resource=uid_str, details={"request_id": req_id_str})
        
        from inventory_app import cache_delete
        cache_delete("dashboard:main")
        
        return True, "Role request cancelled successfully."
    except Exception as e:
        return False, f"Failed to cancel role request: {str(e)}"

def get_user_role_requests(user_id: str) -> list:
    """Returns all role requests submitted by the given user, sorted by created_at desc."""
    uid_str = str(user_id) if user_id else ""
    if not uid_str or not ObjectId.is_valid(uid_str):
        return []
    db = get_db()
    cursor = db.role_requests.find({"user_id": uid_str}).sort("created_at", -1)
    results = []
    for r in cursor:
        r["_id"] = str(r["_id"])
        results.append(r)
    return results

def get_user_pending_role_request(user_id: str) -> dict:
    """Returns any PENDING role request for the given user."""
    uid_str = str(user_id) if user_id else ""
    if not uid_str or not ObjectId.is_valid(uid_str):
        return None
    db = get_db()
    req = db.role_requests.find_one({"user_id": uid_str, "status": "PENDING"})
    if req:
        req["_id"] = str(req["_id"])
    return req

def get_all_pending_role_requests() -> list:
    """Returns list of all PENDING role promotion requests for Admin review."""
    db = get_db()
    requests = list(db.role_requests.find({"status": "PENDING"}).sort("created_at", -1))
    for r in requests:
        r["_id"] = str(r["_id"])
    return requests

def get_all_role_requests() -> list:
    """Returns list of all role promotion requests sorted by created_at desc."""
    db = get_db()
    requests = list(db.role_requests.find().sort("created_at", -1))
    for r in requests:
        r["_id"] = str(r["_id"])
    return requests

def get_role_requests_by_status_count() -> list:
    """Returns a list of status and count for all role requests, used for dashboard visualization."""
    db = get_db()
    pipeline = [
        {"$group": {"_id": "$status", "count": {"$sum": 1}}}
    ]
    try:
        results = list(db.role_requests.aggregate(pipeline))
        return [{"status": r["_id"], "count": r["count"]} for r in results]
    except Exception:
        return []

def process_role_request(request_id: str, action: str, processed_by: str, admin_comment: str = "") -> tuple[bool, str]:
    """Approves or rejects a pending role promotion request with an optional comment."""
    req_id_str = str(request_id) if request_id else ""
    if not req_id_str or not ObjectId.is_valid(req_id_str):
        return False, "Invalid request ID."
    db = get_db()
    try:
        req = db.role_requests.find_one({"_id": ObjectId(req_id_str)})
        if not req:
            return False, "Role request not found."
            
        if req.get("status") != "PENDING":
            return False, "Role request has already been processed."
            
        now = datetime.now(timezone.utc)
        comment_str = admin_comment.strip()
        target_emp = req.get("employee_id", "")
        
        if action == "approve":
            user_id = str(req["user_id"])
            new_role = req.get("requested_role", "inventory_manager")
            
            success, msg = update_user_role(user_id, new_role)
            if not success:
                return False, msg
                
            db.role_requests.update_one(
                {"_id": ObjectId(req_id_str)},
                {"$set": {"status": "APPROVED", "processed_by": processed_by, "admin_comment": comment_str, "updated_at": now}}
            )
            
            from inventory_app.services.audit_service import log_audit
            log_audit("role_request_approved", processed_by, target_resource=user_id, details={"request_id": req_id_str, "admin_comment": comment_str})
            
            from inventory_app import cache_delete
            cache_delete("dashboard:main")
            
            role_label = new_role.replace('_', ' ').title()
            return True, f"Role request approved. User '{target_emp}' is now an {role_label}."
            
        elif action == "reject":
            db.role_requests.update_one(
                {"_id": ObjectId(req_id_str)},
                {"$set": {"status": "REJECTED", "processed_by": processed_by, "admin_comment": comment_str, "updated_at": now}}
            )
            
            from inventory_app.services.audit_service import log_audit
            log_audit("role_request_rejected", processed_by, target_resource=str(req["user_id"]), details={"request_id": req_id_str, "admin_comment": comment_str})
            
            from inventory_app import cache_delete
            cache_delete("dashboard:main")
            
            return True, f"Promotion request for user '{target_emp}' has been rejected."
        else:
            return False, "Invalid action."
    except Exception as e:
        return False, f"Failed to process role request: {str(e)}"


def import_staff_bulk(file_storage, default_password: str = "Staff@123", imported_by: str = "admin") -> tuple[bool, str, dict]:
    """
    Parses and bulk registers staff from an Excel (.xlsx, .xls) or CSV file.
    Expected format: Employee ID, Full Name, Phone No, Email (optional).
    Role is automatically assigned as 'staff'.
    Password defaults to `default_password` (no password column in template).
    Returns (success: bool, message: str, details: dict).
    """
    import io
    import csv
    import re
    from datetime import datetime, timezone
    from werkzeug.security import generate_password_hash
    from inventory_app.database import get_db

    if not file_storage or not file_storage.filename:
        return False, "No file selected. Please choose a valid Excel or CSV file.", {}

    filename = file_storage.filename.lower()
    file_bytes = file_storage.read()
    file_storage.seek(0)

    rows = []
    if filename.endswith(('.xlsx', '.xls')):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
            sheet = wb.active or wb.worksheets[0]
            rows = list(sheet.iter_rows(values_only=True))
        except Exception as e:
            return False, f"Failed to read Excel workbook: {str(e)}", {}
    elif filename.endswith('.csv'):
        try:
            text_content = file_bytes.decode('utf-8-sig', errors='replace')
            csv_reader = csv.reader(io.StringIO(text_content))
            rows = [list(r) for r in csv_reader]
        except Exception as e:
            return False, f"Failed to read CSV file: {str(e)}", {}
    else:
        return False, "Unsupported file format. Please upload an Excel (.xlsx, .xls) or CSV file.", {}

    if not rows or len(rows) < 2:
        return False, "The uploaded file is empty or missing data rows.", {}

    # Identify header row & column mappings
    header = [str(cell or "").strip().lower() for cell in rows[0]]
    col_map = {}
    for idx, col in enumerate(header):
        clean_col = col.replace("_", " ").replace("-", " ").strip()
        if any(k in clean_col for k in ['employee id', 'employeeid', 'empid', 'emp id', 'emp no', 'staff id', 'staff no']) or clean_col == 'id':
            col_map.setdefault('employee_id', idx)
        elif any(k in clean_col for k in ['full name', 'fullname', 'employee name', 'staff name']) or clean_col == 'name':
            col_map.setdefault('name', idx)
        elif any(k in clean_col for k in ['last name', 'lastname', 'surname']):
            col_map.setdefault('surname', idx)
        elif any(k in clean_col for k in ['phone', 'phone no', 'phone number', 'mobile', 'mobile no', 'contact', 'contact no', 'mobile number']):
            col_map.setdefault('phone', idx)
        elif any(k in clean_col for k in ['email', 'mail', 'email id', 'email address']):
            col_map.setdefault('email', idx)
        elif any(k in clean_col for k in ['username', 'user name', 'login']):
            col_map.setdefault('username', idx)

    # Fallback if employee_id column not found but username exists
    if 'employee_id' not in col_map and 'username' in col_map:
        col_map['employee_id'] = col_map['username']

    if 'name' not in col_map and 'employee_id' not in col_map and 'email' not in col_map:
        return False, "Could not detect required columns (Employee ID, Full Name, Phone No) in the file header.", {}

    db = get_db()
    existing_users = list(db.users.find({}, {"email": 1, "employee_id": 1}))
    existing_emails = {u.get("email", "").strip().lower() for u in existing_users if u.get("email")}
    existing_emp_ids = {u.get("employee_id", "").strip().lower() for u in existing_users if u.get("employee_id")}

    batch_emails = set()
    batch_emp_ids = set()
    valid_docs = []
    errors = []
    total_processed = 0

    password_to_use = default_password if (default_password and len(default_password) >= 6) else "Staff@123"
    hashed_password = generate_password_hash(password_to_use)

    for row_idx, row in enumerate(rows[1:], start=2):
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue
        total_processed += 1

        def get_val(key):
            idx = col_map.get(key)
            if idx is not None and idx < len(row) and row[idx] is not None:
                return str(row[idx]).strip()
            return ""

        raw_emp_id = get_val('employee_id')
        raw_name = get_val('name')
        raw_surname = get_val('surname')
        raw_phone = get_val('phone')
        raw_email = get_val('email').lower()

        # 1. Full Name resolution
        full_name = f"{raw_name} {raw_surname}".strip() if raw_surname else raw_name
        if not full_name:
            errors.append(f"Row {row_idx}: Missing employee full name.")
            continue

        # 2. Employee ID resolution
        if not raw_emp_id:
            base_id = "".join(c for c in full_name if c.isalnum()).upper()[:6] or "EMP"
            raw_emp_id = f"{base_id}-{total_processed + 1000}"

        clean_emp_id = raw_emp_id.strip()
        clean_emp_id_lower = clean_emp_id.lower()

        if clean_emp_id_lower in existing_emp_ids or clean_emp_id_lower in batch_emp_ids:
            errors.append(f"Row {row_idx}: Employee ID '{clean_emp_id}' is already registered or duplicated in file.")
            continue

        # 3. Email resolution
        if not raw_email:
            raw_email = f"{re.sub(r'[^a-zA-Z0-9]', '', clean_emp_id).lower()}@stocksetu.local"
        elif "@" not in raw_email or "." not in raw_email:
            errors.append(f"Row {row_idx}: Invalid email address ('{raw_email}').")
            continue

        if raw_email in existing_emails or raw_email in batch_emails:
            errors.append(f"Row {row_idx}: Email '{raw_email}' is already registered or duplicated in file.")
            continue

        now = datetime.now(timezone.utc)
        user_doc = {
            "employee_id": clean_emp_id,
            "employee_id_lower": clean_emp_id_lower,
            "name": full_name or clean_emp_id,
            "name_lower": full_name or clean_emp_id,
            "phone": raw_phone,
            "email": raw_email,
            "email_lower": raw_email,
            "password_hash": hashed_password,
            "role": "staff",
            "is_active": False,
            "force_password_change": True,
            "last_active_at": now,
            "created_at": now,
            "updated_at": now
        }

        valid_docs.append(user_doc)
        batch_emails.add(raw_email)
        batch_emp_ids.add(clean_emp_id_lower)

    if not valid_docs:
        err_detail = "; ".join(errors[:5]) if errors else "No valid staff records found in file."
        return False, f"Import failed: {err_detail}", {"total_rows": total_processed, "imported_count": 0, "errors": errors}

    try:
        result = db.users.insert_many(valid_docs)
        inserted_ids = result.inserted_ids
        inserted_count = len(inserted_ids)

        from inventory_app.services.audit_service import log_audit
        log_audit("bulk_staff_imported", imported_by, details={
            "imported_count": inserted_count,
            "skipped_count": len(errors),
            "total_rows": total_processed
        })

        summary = {
            "total_rows": total_processed,
            "imported_count": inserted_count,
            "role_request_count": 0,
            "skipped_count": len(errors),
            "errors": errors,
            "imported_users": [{"employee_id": d.get("employee_id", ""), "name": d["name"], "phone": d.get("phone", ""), "email": d["email"], "role": d["role"]} for d in valid_docs]
        }
        msg = f"Successfully imported {inserted_count} staff member(s)."
        msg += " All imported accounts have the 'staff' role and must be activated by an Administrator."
        if errors:
            msg += f" {len(errors)} row(s) were skipped due to validation issues."

        return True, msg, summary
    except Exception as e:
        return False, f"Database insertion failed: {str(e)}", {"total_rows": total_processed, "imported_count": 0, "errors": [str(e)]}


def generate_staff_template(file_format: str = "xlsx") -> tuple[io.BytesIO, str, str]:
    """
    Generates a pre-formatted sample Excel/CSV template for bulk staff import.
    Headers: Employee ID, Full Name, Phone No, Email
    Returns (BytesIO_stream, filename, mimetype).
    """
    import io
    if file_format == "csv":
        import csv
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["Employee ID", "Full Name", "Phone No", "Email"])
        writer.writerow(["EMP-1001", "Rahul Sharma", "9876543210", "rahul.sharma@example.com"])
        writer.writerow(["EMP-1002", "Priya Patel", "9876543211", "priya.patel@example.com"])
        writer.writerow(["EMP-1003", "Amit Verma", "9876543212", "amit.verma@example.com"])
        writer.writerow(["EMP-1004", "Neha Gupta", "9876543213", "neha.gupta@example.com"])
        
        mem = io.BytesIO(output.getvalue().encode('utf-8'))
        mem.seek(0)
        return mem, "StockSetu_Staff_Import_Template.csv", "text/csv"

    # Default Excel (.xlsx)
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Staff Import"

    # Headers
    headers = ["Employee ID", "Full Name", "Phone No", "Email"]
    ws.append(headers)

    # Sample rows
    samples = [
        ["EMP-1001", "Rahul Sharma", "9876543210", "rahul.sharma@example.com"],
        ["EMP-1002", "Priya Patel", "9876543211", "priya.patel@example.com"],
        ["EMP-1003", "Amit Verma", "9876543212", "amit.verma@example.com"],
        ["EMP-1004", "Neha Gupta", "9876543213", "neha.gupta@example.com"],
    ]
    for row in samples:
        ws.append(row)

    # Style Header
    header_fill = PatternFill(start_color="0B2545", end_color="0B2545", fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Column widths
    col_widths = [18, 26, 20, 32]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # Add instructions block below
    ws.append([])
    ws.append(["Instructions & Notes:"])
    ws.append(["1. Employee ID and Full Name are required fields."])
    ws.append(["2. Phone No is recommended for staff identification and contact."])
    ws.append(["3. Email is optional (defaults to <EmployeeID>@stocksetu.local if blank)."])
    ws.append(["4. Role is automatically set to 'staff' for all imported members."])
    ws.append(["5. Passwords default to the configured default password (e.g. 'Staff@123') and can be changed in profile settings."])
    ws.append(["6. Staff can log in directly using their Employee ID or Email."])

    note_font = Font(name="Segoe UI", size=9, italic=True, color="64748B")
    for r in range(7, 13):
        cell = ws.cell(row=r, column=1)
        cell.font = note_font

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output, "StockSetu_Staff_Import_Template.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

