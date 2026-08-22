from flask import Blueprint, render_template, request, redirect, url_for, flash, session, current_app
from inventory_app.utils.decorators import login_required, roles_required, csrf_protected

billing_bp = Blueprint('billing', __name__)

# Rate limit state: {ip: [timestamps]}
_rate_limit_state = {}
RATE_LIMIT_WINDOW = 60  # seconds
RATE_LIMIT_MAX = 60     # max requests per window
_RATE_LIMIT_MAX_IPS = 500  # max tracked IPs to prevent OOM


def _check_rate_limit():
    """Simple in-memory sliding-window rate limiter for POST routes."""
    import time
    from flask import request as req, abort
    if req.method != 'POST':
        return
    ip = req.remote_addr or 'unknown'
    now = time.time()
    window_start = now - RATE_LIMIT_WINDOW

    # Evict stale entries and cap dict size to prevent OOM
    if len(_rate_limit_state) > _RATE_LIMIT_MAX_IPS:
        stale_keys = [k for k, v in _rate_limit_state.items()
                      if not v or v[-1] < window_start]
        for k in stale_keys[:len(stale_keys) // 2]:
            _rate_limit_state.pop(k, None)

    hits = _rate_limit_state.get(ip, [])
    hits = [t for t in hits if t > window_start]
    if len(hits) >= RATE_LIMIT_MAX:
        abort(429)
    hits.append(now)
    _rate_limit_state[ip] = hits


@billing_bp.before_request
def _billing_rate_limit():
    _check_rate_limit()


def _limiter():
    """Returns the Flask-Limiter instance if available, else None."""
    return current_app.extensions.get('limiter')


def get_category_theme_idx(category_name):
    """Deterministic hash for category name to map to one of 16 theme colors."""
    if not category_name:
        return 0
    h = 0
    for char in str(category_name).strip().lower():
        h = (h * 31 + ord(char)) % 16
    return h


@billing_bp.route('/billing')
@login_required
@roles_required('admin', 'inventory_manager', 'staff')
def pos():
    """POS quick-billing screen: search products, build cart, create GST invoice."""
    from inventory_app.services.product_service import search_products, get_distinct_categories
    from inventory_app.services.billing_service import get_active_employees_for_billing
    query = request.args.get('q', '').strip()
    category = request.args.get('category', '').strip()
    # Load only first page for initial render; subsequent searches use API
    products, _ = search_products(
        query=query, category=category, is_active=True,
        sort_by="product_name", sort_dir=1,
        page=1, per_page=50, return_total=True
    )
    categories = get_distinct_categories()
    employees = get_active_employees_for_billing()
    return render_template(
        'billing/pos.html',
        products=products,
        categories=categories,
        employees=employees,
        current_query=query,
        current_category=category,
        get_category_theme_idx=get_category_theme_idx
    )


@billing_bp.route('/billing/create', methods=['POST'])
@login_required
@roles_required('admin', 'inventory_manager', 'staff')
@csrf_protected
def create_bill():
    """Handles POS bill submission. Rate limited to 30 creates/min per user."""
    from inventory_app.services.billing_service import create_bill
    is_employee_purchase = request.form.get('is_employee_purchase') in ('1', 'true', 'True', 'on')
    customer_data = {
        'customer_name': request.form.get('customer_name', ''),
        'customer_phone': request.form.get('customer_phone', ''),
        'customer_gstin': request.form.get('customer_gstin', ''),
        'is_employee_purchase': is_employee_purchase,
        'buyer_employee_id': request.form.get('buyer_employee_id', '').strip(),
        'staff_discount_percent': request.form.get('staff_discount_percent', '10'),
        'payment_method': request.form.get('payment_method', 'CASH'),
        'discount_percent': request.form.get('discount_percent', '0'),
        'due_date': request.form.get('due_date', '') or None,
    }

    items = []
    names = request.form.getlist('item_name[]')
    quantities = request.form.getlist('item_quantity[]')
    discounts = request.form.getlist('item_discount[]')
    free_flags = request.form.getlist('item_free[]')
    for i, name in enumerate(names):
        if not name.strip():
            continue
        item = {
            'product_name': name,
            'quantity': quantities[i] if i < len(quantities) else '1',
        }
        if i < len(discounts) and discounts[i]:
            item['line_discount_percent'] = discounts[i]
        if i < len(free_flags) and free_flags[i] == '1':
            item['is_free'] = True
        items.append(item)

    charges = {
        'shipping_charge': request.form.get('shipping_charge', '0'),
        'packing_charge': request.form.get('packing_charge', '0'),
    }

    # Parse payment splits if provided
    payment_splits = []
    split_methods = request.form.getlist('payment_split_method[]')
    split_amounts = request.form.getlist('payment_split_amount[]')
    split_references = request.form.getlist('payment_split_reference[]')
    if split_methods and split_amounts:
        for i, method in enumerate(split_methods):
            if method.strip():
                amount = split_amounts[i] if i < len(split_amounts) else '0'
                reference = split_references[i] if i < len(split_references) else ''
                payment_splits.append({
                    'method': method.strip().upper(),
                    'amount': amount,
                    'reference': reference.strip() if reference else '',
                })

    performed_by = session.get('employee_id') or session.get('name') or 'System'
    success, msg, bill = create_bill(
        customer_data, items, performed_by=performed_by,
        charges=charges,
        payment_splits=payment_splits if payment_splits else None,
    )

    if success:
        flash(msg, "success")
        return redirect(url_for('billing.view_bill', bill_id=bill['_id']))

    flash(msg, "danger")
    return redirect(url_for('billing.pos'))


@billing_bp.route('/billing/bills')
@login_required
def list_bills():
    from inventory_app.services.billing_service import get_bills
    from inventory_app.utils.pagination import Pagination
    search = request.args.get('q', '').strip()
    status = request.args.get('status', '').strip()
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 25, type=int)
    
    bills, total_count = get_bills(
        search=search,
        payment_status=status,
        page=page,
        per_page=per_page,
        return_total=True
    )
    pagination = Pagination(page=page, per_page=per_page, total=total_count)
    return render_template(
        'billing/bills.html',
        bills=bills,
        pagination=pagination,
        current_search=search,
        current_status=status
    )


@billing_bp.route('/billing/bills/<bill_id>')
@login_required
def view_bill(bill_id):
    from inventory_app.services.billing_service import get_bill_by_id, get_bill_audit_history
    bill = get_bill_by_id(bill_id)
    if not bill:
        flash("Bill not found.", "warning")
        return redirect(url_for('billing.list_bills'))
    audit_history = get_bill_audit_history(bill.get("bill_number", ""))
    return render_template(
        'billing/bill_detail.html',
        bill=bill,
        audit_history=audit_history,
    )


@billing_bp.route('/billing/bills/<bill_id>/refund', methods=['POST'])
@login_required
@roles_required('admin', 'inventory_manager')
@csrf_protected
def refund_bill(bill_id):
    from inventory_app.services.billing_service import refund_bill
    reason = request.form.get('reason', 'Customer Refund')
    username = session.get('employee_id', 'Unknown')
    success, msg = refund_bill(bill_id, reason, username)
    if success:
        flash(msg, "success")
    else:
        flash(msg, "danger")
    return redirect(url_for('billing.view_bill', bill_id=bill_id))


@billing_bp.route('/billing/bills/<bill_id>/refund-items', methods=['POST'])
@login_required
@roles_required('admin', 'inventory_manager')
@csrf_protected
def refund_items(bill_id):
    """Refund specific line(s) of a bill."""
    from inventory_app.services.billing_service import refund_bill_lines
    line_indices = request.form.getlist('line_index[]')
    reason = request.form.get('reason', 'Line refund')
    username = session.get('employee_id', 'Unknown')
    success, msg = refund_bill_lines(bill_id, line_indices, reason, username)
    flash(msg, "success" if success else "danger")
    return redirect(url_for('billing.view_bill', bill_id=bill_id))


@billing_bp.route('/billing/bills/<bill_id>/pay', methods=['POST'])
@login_required
@roles_required('admin', 'inventory_manager', 'staff')
@csrf_protected
def record_payment(bill_id):
    """Record a payment against a bill."""
    from inventory_app.services.billing_service import record_bill_payment
    try:
        amount = float(request.form.get('amount', 0))
    except (ValueError, TypeError):
        amount = 0
    method = request.form.get('method', 'CASH')
    reference = request.form.get('reference', '')
    username = session.get('employee_id', 'Unknown')
    success, msg = record_bill_payment(bill_id, amount, method, reference, username)
    flash(msg, "success" if success else "danger")
    return redirect(url_for('billing.view_bill', bill_id=bill_id))


@billing_bp.route('/billing/bills/<bill_id>/edit', methods=['POST'])
@login_required
@roles_required('admin', 'inventory_manager')
@csrf_protected
def edit_bill(bill_id):
    """Edit a bill's line items and charges."""
    from inventory_app.services.billing_service import edit_bill
    customer_data = {
        'customer_name': request.form.get('customer_name', ''),
        'customer_phone': request.form.get('customer_phone', ''),
        'customer_gstin': request.form.get('customer_gstin', ''),
        'payment_method': request.form.get('payment_method', 'CASH'),
        'discount_percent': request.form.get('discount_percent', '0'),
        'due_date': request.form.get('due_date', '') or None,
    }

    items = []
    names = request.form.getlist('item_name[]')
    quantities = request.form.getlist('item_quantity[]')
    discounts = request.form.getlist('item_discount[]')
    free_flags = request.form.getlist('item_free[]')
    for i, name in enumerate(names):
        if not name.strip():
            continue
        item = {
            'product_name': name,
            'quantity': quantities[i] if i < len(quantities) else '1',
        }
        if i < len(discounts) and discounts[i]:
            item['line_discount_percent'] = discounts[i]
        if i < len(free_flags) and free_flags[i] == '1':
            item['is_free'] = True
        items.append(item)

    charges = {
        'shipping_charge': request.form.get('shipping_charge', '0'),
        'packing_charge': request.form.get('packing_charge', '0'),
    }

    username = session.get('employee_id', 'Unknown')
    success, msg, bill = edit_bill(bill_id, items, charges, customer_data, username)
    flash(msg, "success" if success else "danger")
    return redirect(url_for('billing.view_bill', bill_id=bill_id))


@billing_bp.route('/billing/reconciliation')
@login_required
@roles_required('admin')
def reconciliation():
    """Admin reconciliation report: flags anomalies across bills, stock, and audit."""
    from inventory_app.services.billing_service import get_reconciliation_report
    anomalies = get_reconciliation_report()
    return render_template('billing/reconciliation.html', anomalies=anomalies)


@billing_bp.route('/billing/sales')
@login_required
@roles_required('admin', 'inventory_manager', 'staff')
def sales_analytics():
    """Sales Analytics & Staff Performance executive dashboard."""
    from inventory_app.services.billing_service import get_sales_analytics

    date_preset = request.args.get('range', '30d').strip()
    start_date = request.args.get('start_date', '').strip()
    end_date = request.args.get('end_date', '').strip()
    cashier = request.args.get('cashier', '').strip()

    analytics = get_sales_analytics(
        date_preset=date_preset,
        start_date=start_date,
        end_date=end_date,
        cashier=cashier
    )

    return render_template(
        'billing/sales.html',
        analytics=analytics,
        current_range=date_preset,
        current_cashier=cashier,
        start_date=start_date,
        end_date=end_date
    )


@billing_bp.route('/billing/sales/export')
@login_required
@roles_required('admin', 'inventory_manager', 'staff')
def export_sales():
    """Exports sales analytics report as CSV."""
    import csv
    import io
    from flask import Response
    from inventory_app.services.billing_service import get_sales_analytics

    date_preset = request.args.get('range', '30d').strip()
    start_date = request.args.get('start_date', '').strip()
    end_date = request.args.get('end_date', '').strip()
    cashier = request.args.get('cashier', '').strip()

    analytics = get_sales_analytics(date_preset, start_date, end_date, cashier)
    
    output = io.StringIO()
    writer = csv.writer(output)
    
    # 1. Summary
    writer.writerow(["STOCKSETU SALES ANALYTICS REPORT"])
    writer.writerow(["Date Range", date_preset, "Start", analytics.get("start_date"), "End", analytics.get("end_date")])
    writer.writerow([])
    
    # 2. KPIs
    kpi = analytics.get("kpi", {})
    writer.writerow(["KEY PERFORMANCE INDICATORS"])
    writer.writerow(["Total Gross Sales (INR)", kpi.get("total_sales", 0)])
    writer.writerow(["Total Invoices", kpi.get("total_bills", 0)])
    writer.writerow(["Average Bill Value (INR)", kpi.get("avg_bill_value", 0)])
    writer.writerow(["Total Tax Collected (INR)", kpi.get("total_tax", 0)])
    writer.writerow(["Total Discounts Given (INR)", kpi.get("total_discount", 0)])
    writer.writerow(["Total Refunds (INR)", kpi.get("total_refunded", 0)])
    writer.writerow([])
    
    # 3. Staff Leaderboard
    writer.writerow(["STAFF / CASHIER PERFORMANCE"])
    writer.writerow(["Cashier Username", "Total Bills", "Total Sales (INR)", "Avg Order (INR)", "Cash Sales", "UPI Sales", "Card Sales", "Credit Sales"])
    for s in analytics.get("staff_leaderboard", []):
        writer.writerow([
            s.get("cashier"),
            s.get("bill_count"),
            s.get("total_sales"),
            s.get("avg_sale"),
            s.get("cash_sales"),
            s.get("upi_sales"),
            s.get("card_sales"),
            s.get("credit_sales")
        ])
    writer.writerow([])

    # 4. Top Selling Products
    writer.writerow(["TOP SELLING PRODUCTS"])
    writer.writerow(["Product Name", "Quantity Sold", "Revenue (INR)", "Bill Appearances"])
    for p in analytics.get("top_products", []):
        writer.writerow([p.get("product_name"), p.get("quantity_sold"), p.get("revenue"), p.get("bill_count")])

    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment;filename=sales_report_{date_preset}.csv"}
    )


@billing_bp.route('/billing/sales/export/excel')
@login_required
@roles_required('admin', 'inventory_manager', 'staff')
def export_sales_excel():
    """Exports sales analytics as a rich Excel workbook with Summary, Staff Performance, and Full Data sheets."""
    import io
    from datetime import datetime
    from flask import send_file
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from inventory_app.services.billing_service import get_sales_analytics

    date_preset = request.args.get('range', '30d').strip()
    start_date = request.args.get('start_date', '').strip()
    end_date = request.args.get('end_date', '').strip()
    cashier = request.args.get('cashier', '').strip()
    analytics = get_sales_analytics(date_preset, start_date, end_date, cashier)
    kpi = analytics.get("kpi", {})
    staff = analytics.get("staff_leaderboard", [])
    products = analytics.get("top_products", [])

    # ── Reusable Styles ──
    BRAND = '0B3D6E'
    WHITE = 'FFFFFF'
    LIGHT_BG = 'F1F5F9'
    GREEN = '16A34A'
    RED = 'DC2626'
    AMBER = 'D97706'
    PURPLE = '7C3AED'
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'), right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'), bottom=Side(style='thin', color='CBD5E1'),
    )
    header_font = Font(name='Calibri', bold=True, size=10, color=WHITE)
    header_fill = PatternFill(start_color=BRAND, end_color=BRAND, fill_type='solid')
    data_font = Font(name='Calibri', size=10)
    currency_fmt = '₹#,##0.00'
    center = Alignment(horizontal='center', vertical='center')
    right = Alignment(horizontal='right', vertical='center')

    def _apply_header(ws, row, cols):
        for c in range(1, cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = thin_border

    def _apply_data_row(ws, row, cols, stripe=False):
        for c in range(1, cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = data_font
            cell.border = thin_border
            if stripe:
                cell.fill = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')

    def _auto_width(ws, min_w=10, max_w=35):
        for col_idx in range(1, ws.max_column + 1):
            best = min_w
            for row_idx in range(1, min(ws.max_row + 1, 200)):
                cell = ws.cell(row=row_idx, column=col_idx)
                try:
                    if cell.value:
                        best = max(best, min(len(str(cell.value)) + 3, max_w))
                except:
                    pass
            ws.column_dimensions[get_column_letter(col_idx)].width = best

    def _currency_cells(ws, row, start_col, end_col):
        for c in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=c)
            cell.number_format = currency_fmt
            cell.alignment = right

    wb = Workbook()

    # ════════════════════════════════════════════════════════════
    # SHEET 1: Summary
    # ════════════════════════════════════════════════════════════
    summ_ws = wb.active
    summ_ws.title = "Summary"
    summ_ws.sheet_properties.tabColor = '16A34A'

    summ_ws.merge_cells('A1:E1')
    st = summ_ws.cell(row=1, column=1, value="STOCKSETU — SALES SUMMARY")
    st.font = Font(name='Calibri', bold=True, size=16, color=WHITE)
    st.fill = PatternFill(start_color=BRAND, end_color=BRAND, fill_type='solid')
    st.alignment = Alignment(horizontal='center', vertical='center')
    summ_ws.row_dimensions[1].height = 36

    summ_ws.merge_cells('A2:E2')
    ss = summ_ws.cell(row=2, column=1, value=f"Period: {date_preset}  |  {analytics.get('start_date','')} → {analytics.get('end_date','')}")
    ss.font = Font(name='Calibri', size=9, color='64748B')
    ss.fill = PatternFill(start_color=LIGHT_BG, end_color=LIGHT_BG, fill_type='solid')
    ss.alignment = Alignment(horizontal='center', vertical='center')
    summ_ws.row_dimensions[2].height = 20

    # KPIs in 2-column layout
    summ_data = [
        ("Total Gross Sales", kpi.get("total_sales", 0), currency_fmt, GREEN),
        ("Net Collected", kpi.get("total_paid", 0), currency_fmt, GREEN),
        ("Pending / Due", kpi.get("total_due", 0), currency_fmt, RED),
        ("Total Tax Collected", kpi.get("total_tax", 0), currency_fmt, AMBER),
        ("Total Discounts", kpi.get("total_discount", 0), currency_fmt, RED),
        ("Total Refunds", kpi.get("total_refunded", 0), currency_fmt, RED),
        ("Total Invoices", kpi.get("total_bills", 0), '#,##0', BRAND),
        ("Average Bill Value", kpi.get("avg_bill_value", 0), currency_fmt, PURPLE),
    ]

    r = 4
    for i in range(0, len(summ_data), 2):
        lbl, val, fmt, color = summ_data[i]
        summ_ws.cell(row=r, column=1, value=lbl).font = Font(name='Calibri', bold=True, size=10, color='334155')
        summ_ws.cell(row=r, column=1).border = Border(bottom=Side(style='thin', color='E2E8F0'))
        vc = summ_ws.cell(row=r, column=2, value=val)
        vc.font = Font(name='Calibri', bold=True, size=12, color=color)
        vc.number_format = fmt
        vc.alignment = right
        vc.border = Border(bottom=Side(style='thin', color='E2E8F0'))
        if i + 1 < len(summ_data):
            lbl2, val2, fmt2, color2 = summ_data[i + 1]
            summ_ws.cell(row=r, column=4, value=lbl2).font = Font(name='Calibri', bold=True, size=10, color='334155')
            summ_ws.cell(row=r, column=4).border = Border(bottom=Side(style='thin', color='E2E8F0'))
            vc2 = summ_ws.cell(row=r, column=5, value=val2)
            vc2.font = Font(name='Calibri', bold=True, size=12, color=color2)
            vc2.number_format = fmt2
            vc2.alignment = right
            vc2.border = Border(bottom=Side(style='thin', color='E2E8F0'))
        r += 1

    r += 1
    # Top Staff mini-table
    summ_ws.merge_cells(f'A{r}:E{r}')
    sh = summ_ws.cell(row=r, column=1, value="TOP PERFORMERS")
    sh.font = Font(name='Calibri', bold=True, size=11, color=BRAND)
    sh.fill = PatternFill(start_color='DBEAFE', end_color='DBEAFE', fill_type='solid')
    sh.border = Border(bottom=Side(style='medium', color=BRAND))
    r += 1

    for ci, h in enumerate(["Rank", "Staff", "Bills", "Total Sales", "Avg Ticket"], 1):
        summ_ws.cell(row=r, column=ci, value=h)
    _apply_header(summ_ws, r, 5)
    r += 1

    for i, s in enumerate(staff[:5], 1):
        summ_ws.cell(row=r, column=1, value=i).alignment = center
        summ_ws.cell(row=r, column=2, value=f"{s.get('name','')} (@{s.get('cashier','')})")
        summ_ws.cell(row=r, column=3, value=s.get("bill_count", 0)).alignment = center
        summ_ws.cell(row=r, column=4, value=s.get("total_sales", 0))
        summ_ws.cell(row=r, column=5, value=s.get("avg_sale", 0))
        _currency_cells(summ_ws, r, 4, 5)
        _apply_data_row(summ_ws, r, 5, stripe=(i % 2 == 0))
        r += 1

    r += 1
    # Top Products mini-table
    summ_ws.merge_cells(f'A{r}:E{r}')
    ph = summ_ws.cell(row=r, column=1, value="TOP PRODUCTS")
    ph.font = Font(name='Calibri', bold=True, size=11, color=BRAND)
    ph.fill = PatternFill(start_color='DBEAFE', end_color='DBEAFE', fill_type='solid')
    ph.border = Border(bottom=Side(style='medium', color=BRAND))
    r += 1

    for ci, h in enumerate(["#", "Product", "Qty Sold", "Revenue", "Invoices"], 1):
        summ_ws.cell(row=r, column=ci, value=h)
    _apply_header(summ_ws, r, 5)
    r += 1

    for i, p in enumerate(products[:5], 1):
        summ_ws.cell(row=r, column=1, value=i).alignment = center
        summ_ws.cell(row=r, column=2, value=p.get("product_name", ""))
        summ_ws.cell(row=r, column=3, value=int(p.get("quantity_sold", 0))).alignment = center
        summ_ws.cell(row=r, column=4, value=p.get("revenue", 0))
        summ_ws.cell(row=r, column=5, value=p.get("bill_count", 0)).alignment = center
        _currency_cells(summ_ws, r, 4, 4)
        _apply_data_row(summ_ws, r, 5, stripe=(i % 2 == 0))
        r += 1

    _auto_width(summ_ws, min_w=10, max_w=35)
    summ_ws.sheet_view.showGridLines = False

    # ════════════════════════════════════════════════════════════
    # SHEET 2: Staff Performance
    # ════════════════════════════════════════════════════════════
    staff_ws = wb.create_sheet("Staff Performance")
    staff_ws.sheet_properties.tabColor = 'F59E0B'

    staff_ws.merge_cells('A1:H1')
    sh = staff_ws.cell(row=1, column=1, value="STAFF / CASHIER PERFORMANCE LEADERBOARD")
    sh.font = Font(name='Calibri', bold=True, size=13, color=WHITE)
    sh.fill = PatternFill(start_color=BRAND, end_color=BRAND, fill_type='solid')
    sh.alignment = center
    staff_ws.row_dimensions[1].height = 30

    headers = ["Rank", "Username", "Full Name", "Role", "Bills", "Total Sales", "Avg Ticket", "Cash", "Digital"]
    for ci, h in enumerate(headers, 1):
        staff_ws.cell(row=3, column=ci, value=h)
    _apply_header(staff_ws, 3, len(headers))

    for i, s in enumerate(staff, 1):
        r = i + 3
        staff_ws.cell(row=r, column=1, value=i).alignment = center
        staff_ws.cell(row=r, column=2, value=s.get("cashier", ""))
        staff_ws.cell(row=r, column=3, value=s.get("name", ""))
        role = s.get("role", "staff")
        role_cell = staff_ws.cell(row=r, column=4, value=role.replace("_", " ").title())
        if role == "admin":
            role_cell.font = Font(name='Calibri', bold=True, size=10, color=RED)
        elif role == "inventory_manager":
            role_cell.font = Font(name='Calibri', bold=True, size=10, color=AMBER)
        staff_ws.cell(row=r, column=5, value=s.get("bill_count", 0)).alignment = center
        staff_ws.cell(row=r, column=6, value=s.get("total_sales", 0))
        staff_ws.cell(row=r, column=7, value=s.get("avg_sale", 0))
        staff_ws.cell(row=r, column=8, value=s.get("cash_sales", 0))
        staff_ws.cell(row=r, column=9, value=s.get("upi_sales", 0) + s.get("card_sales", 0))
        _currency_cells(staff_ws, r, 6, 9)
        _apply_data_row(staff_ws, r, len(headers), stripe=(i % 2 == 0))

        rank_cell = staff_ws.cell(row=r, column=1)
        if i == 1:
            rank_cell.fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
            rank_cell.font = Font(name='Calibri', bold=True, size=11, color=AMBER)
        elif i == 2:
            rank_cell.fill = PatternFill(start_color='E0E7FF', end_color='E0E7FF', fill_type='solid')
            rank_cell.font = Font(name='Calibri', bold=True, size=11, color='4338CA')
        elif i == 3:
            rank_cell.fill = PatternFill(start_color='FCE7F3', end_color='FCE7F3', fill_type='solid')
            rank_cell.font = Font(name='Calibri', bold=True, size=11, color='9D174D')

    if staff:
        staff_ws.auto_filter.ref = f"A3:{get_column_letter(len(headers))}{3 + len(staff)}"
        staff_ws.freeze_panes = "A4"
    _auto_width(staff_ws, min_w=10, max_w=30)

    # ════════════════════════════════════════════════════════════
    # SHEET 3: Full Data (raw data for filtering)
    # ════════════════════════════════════════════════════════════
    full_ws = wb.create_sheet("Full Data")
    full_ws.sheet_properties.tabColor = '64748B'

    full_ws.merge_cells('A1:J1')
    fdh = full_ws.cell(row=1, column=1, value="ALL DATA — FILTER & SORT AS NEEDED")
    fdh.font = Font(name='Calibri', bold=True, size=13, color=WHITE)
    fdh.fill = PatternFill(start_color=BRAND, end_color=BRAND, fill_type='solid')
    fdh.alignment = center
    full_ws.row_dimensions[1].height = 28

    # Staff data
    full_ws.merge_cells('A3:J3')
    fs = full_ws.cell(row=3, column=1, value="STAFF PERFORMANCE DATA")
    fs.font = Font(name='Calibri', bold=True, size=11, color=BRAND)
    fs.fill = PatternFill(start_color='DBEAFE', end_color='DBEAFE', fill_type='solid')
    fs.border = Border(bottom=Side(style='medium', color=BRAND))

    staff_full_headers = ["Username", "Full Name", "Role", "Bills", "Total Sales", "Avg Sale", "Cash", "UPI", "Card", "Digital Total"]
    for ci, h in enumerate(staff_full_headers, 1):
        full_ws.cell(row=4, column=ci, value=h)
    _apply_header(full_ws, 4, len(staff_full_headers))

    for i, s in enumerate(staff, 5):
        full_ws.cell(row=i, column=1, value=s.get("cashier", ""))
        full_ws.cell(row=i, column=2, value=s.get("name", ""))
        full_ws.cell(row=i, column=3, value=s.get("role", "staff").replace("_", " ").title())
        full_ws.cell(row=i, column=4, value=s.get("bill_count", 0)).alignment = center
        full_ws.cell(row=i, column=5, value=s.get("total_sales", 0))
        full_ws.cell(row=i, column=6, value=s.get("avg_sale", 0))
        full_ws.cell(row=i, column=7, value=s.get("cash_sales", 0))
        full_ws.cell(row=i, column=8, value=s.get("upi_sales", 0))
        full_ws.cell(row=i, column=9, value=s.get("card_sales", 0))
        full_ws.cell(row=i, column=10, value=s.get("upi_sales", 0) + s.get("card_sales", 0))
        _currency_cells(full_ws, i, 5, 10)
        _apply_data_row(full_ws, i, len(staff_full_headers), stripe=(i % 2 == 0))

    staff_end_row = 4 + len(staff) if staff else 4

    # Products section
    prod_start = staff_end_row + 3
    full_ws.merge_cells(f'A{prod_start}:E{prod_start}')
    ps = full_ws.cell(row=prod_start, column=1, value="PRODUCT PERFORMANCE DATA")
    ps.font = Font(name='Calibri', bold=True, size=11, color=BRAND)
    ps.fill = PatternFill(start_color='DBEAFE', end_color='DBEAFE', fill_type='solid')
    ps.border = Border(bottom=Side(style='medium', color=BRAND))

    prod_full_headers = ["Product Name", "Qty Sold", "Revenue", "Invoices", "Avg per Invoice"]
    for ci, h in enumerate(prod_full_headers, 1):
        full_ws.cell(row=prod_start + 1, column=ci, value=h)
    _apply_header(full_ws, prod_start + 1, len(prod_full_headers))

    for i, p in enumerate(products, prod_start + 2):
        full_ws.cell(row=i, column=1, value=p.get("product_name", ""))
        full_ws.cell(row=i, column=2, value=int(p.get("quantity_sold", 0))).alignment = center
        full_ws.cell(row=i, column=3, value=p.get("revenue", 0))
        full_ws.cell(row=i, column=4, value=p.get("bill_count", 0)).alignment = center
        avg = p.get("revenue", 0) / p.get("bill_count", 1) if p.get("bill_count", 0) > 0 else 0
        full_ws.cell(row=i, column=5, value=avg)
        _currency_cells(full_ws, i, 3, 5)
        _apply_data_row(full_ws, i, len(prod_full_headers), stripe=(i % 2 == 0))

    if staff:
        full_ws.auto_filter.ref = f"A4:{get_column_letter(len(staff_full_headers))}{staff_end_row}"
    full_ws.freeze_panes = "A5"
    _auto_width(full_ws, min_w=10, max_w=35)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"sales_report_{date_preset}.xlsx"
    )


