from flask import Blueprint, render_template, request, redirect, url_for, flash, session
from inventory_app.services.product_service import search_products, get_product_by_name, create_product, normalize_product_name, validate_product_data, calculate_stock_status, invalidate_product_cache
from inventory_app.services.inventory_service import (
    stock_in, stock_out, stock_adjust, get_all_transactions
)
from inventory_app.utils.decorators import login_required, roles_required, csrf_protected
from inventory_app.utils.validators import generate_csrf_token
from inventory_app.database import get_db
from inventory_app.services.audit_service import log_audit
from datetime import datetime, timezone
import re

inventory_bp = Blueprint('inventory', __name__)

@inventory_bp.route('/inventory/stock-in', methods=['GET', 'POST'])
@inventory_bp.route('/products/<product_name>/stock-in', methods=['GET', 'POST'])
@login_required
@roles_required('admin', 'inventory_manager', 'staff')
@csrf_protected
def handle_stock_in(product_name=None):
    if request.method == 'POST':
        p_name = request.form.get('product_name') or product_name
        try:
            quantity = float(request.form.get('quantity', 0))
        except (ValueError, TypeError):
            flash("Quantity must be a valid number.", "danger")
            return redirect(request.referrer or url_for('products.list_products'))
            
        reason = request.form.get('reason', '')
        username = session.get('username', 'System')
        
        success, msg, prod = stock_in(p_name, quantity, reason, performed_by=username)
        if success:
            flash(msg, "success")
            return redirect(url_for('products.view_product', product_name=prod['product_name']))
        else:
            flash(msg, "danger")
            
    products = search_products(is_active=True)
    selected_product = get_product_by_name(product_name) if product_name else None
    return render_template('inventory/stock_in.html', products=products, selected_product=selected_product)

@inventory_bp.route('/inventory/stock-out', methods=['GET', 'POST'])
@inventory_bp.route('/products/<product_name>/stock-out', methods=['GET', 'POST'])
@login_required
@roles_required('admin', 'inventory_manager', 'staff')
@csrf_protected
def handle_stock_out(product_name=None):
    if request.method == 'POST':
        p_name = request.form.get('product_name') or product_name
        try:
            quantity = float(request.form.get('quantity', 0))
        except (ValueError, TypeError):
            flash("Quantity must be a valid number.", "danger")
            return redirect(request.referrer or url_for('products.list_products'))
            
        reason = request.form.get('reason', '')
        username = session.get('username', 'System')
        
        success, msg, prod = stock_out(p_name, quantity, reason, performed_by=username)
        if success:
            flash(msg, "success")
            return redirect(url_for('products.view_product', product_name=prod['product_name']))
        else:
            flash(msg, "danger")
            
    products = search_products(is_active=True)
    selected_product = get_product_by_name(product_name) if product_name else None
    return render_template('inventory/stock_out.html', products=products, selected_product=selected_product)

@inventory_bp.route('/inventory/adjust', methods=['GET', 'POST'])
@inventory_bp.route('/products/<product_name>/adjust', methods=['GET', 'POST'])
@login_required
@roles_required('admin', 'inventory_manager')
@csrf_protected
def handle_adjust(product_name=None):
    if request.method == 'POST':
        p_name = request.form.get('product_name') or product_name
        try:
            target_quantity = float(request.form.get('target_quantity', 0))
        except (ValueError, TypeError):
            flash("Target quantity must be a valid number.", "danger")
            return redirect(request.referrer or url_for('products.list_products'))
            
        reason = request.form.get('reason', '')
        username = session.get('username', 'System')
        
        success, msg, prod = stock_adjust(p_name, target_quantity, reason, performed_by=username)
        if success:
            flash(msg, "success")
            return redirect(url_for('products.view_product', product_name=prod['product_name']))
        else:
            flash(msg, "danger")
            
    products = search_products(is_active=True)
    selected_product = get_product_by_name(product_name) if product_name else None
    return render_template('inventory/adjust.html', products=products, selected_product=selected_product)

@inventory_bp.route('/inventory/bulk-stock-in', methods=['GET', 'POST'])
@login_required
@roles_required('admin', 'inventory_manager')
@csrf_protected
def bulk_stock_in():
    from inventory_app.services.import_service import parse_supplier_bill

    if request.method == 'POST':
        file = request.files.get('bill_file')
        if not file or not file.filename:
            flash("Please select a file to upload.", "danger")
            return redirect(url_for('inventory.bulk_stock_in'))

        ok, err, items = parse_supplier_bill(file)
        if not ok:
            flash(err, "danger")
            return redirect(url_for('inventory.bulk_stock_in'))

        products = search_products(is_active=True, limit=500)
        product_names = [p['product_name'] for p in products]
        product_map = {}
        for p in products:
            key = p['product_name'].strip().lower()
            if key not in product_map:
                product_map[key] = p['product_name']

        for item in items:
            matched = product_map.get(item['item_name'].strip().lower(), None)
            item['is_new'] = matched is None
            item['matched_name'] = matched or ''

        session['bulk_import_items'] = [
            {'item_name': item['item_name'], 'quantity': item['quantity']}
            for item in items
        ]
        return render_template(
            'inventory/bulk_stock_in.html',
            items=items,
            product_names=product_names,
            step='map',
            csrf_value=generate_csrf_token(),
        )

    return render_template('inventory/bulk_stock_in.html', step='upload')

@inventory_bp.route('/inventory/bulk-stock-in/confirm', methods=['POST'])
@login_required
@roles_required('admin', 'inventory_manager')
@csrf_protected
def bulk_stock_in_confirm():
    mapping = request.form.getlist('mapping[]')
    item_names = request.form.getlist('item_name[]')
    item_qtys = request.form.getlist('item_qty[]')
    item_units = request.form.getlist('item_unit[]')
    item_hsns = request.form.getlist('item_hsn[]')
    new_categories = request.form.getlist('new_category[]')
    new_prices = request.form.getlist('new_price[]')
    new_gsts = request.form.getlist('new_gst[]')
    new_descs = request.form.getlist('new_desc[]')
    new_units = request.form.getlist('new_unit[]')
    new_hsns = request.form.getlist('new_hsn[]')
    reason = request.form.get('reason', 'Bulk import from supplier bill')
    username = session.get('username', 'System')

    items = session.pop('bulk_import_items', None)
    if not items:
        items = []
        for i, name in enumerate(item_names):
            try:
                qty = float(item_qtys[i]) if i < len(item_qtys) else 0
            except (ValueError, TypeError):
                qty = 0
            items.append({
                'item_name': name,
                'quantity': qty,
                'unit': item_units[i] if i < len(item_units) else '',
                'hsn_code': item_hsns[i] if i < len(item_hsns) else '',
            })

    if not items:
        flash("No items to import. Please upload the bill again.", "danger")
        return redirect(url_for('inventory.bulk_stock_in'))

    success_count = 0
    created_count = 0
    errors = []
    new_i = 0
    now = datetime.now(timezone.utc)

    new_products = []
    new_transactions = []
    existing_updates = []

    for i, item in enumerate(items):
        product_name = mapping[i] if i < len(mapping) else ""
        if not product_name:
            continue

        qty = float(item.get('quantity', 0))
        if qty <= 0:
            continue

        if product_name == '__new__':
            category = new_categories[new_i] if new_i < len(new_categories) else ''
            price = float(new_prices[new_i]) if new_i < len(new_prices) else 0
            gst = float(new_gsts[new_i]) if new_i < len(new_gsts) else 0
            desc = new_descs[new_i] if new_i < len(new_descs) else ''
            unit_from_form = new_units[new_i] if new_i < len(new_units) else ''
            hsn_from_form = new_hsns[new_i] if new_i < len(new_hsns) else ''
            new_i += 1

            if not category.strip():
                errors.append(f"'{item['item_name']}' — Category is required")
                continue

            canonical = normalize_product_name(item['item_name'])

            product_doc = {
                "product_name": canonical,
                "category": category.strip(),
                "description": desc.strip(),
                "quantity": qty,
                "unit": (unit_from_form or item.get('unit', '') or '').strip(),
                "price": price,
                "gst_rate": gst,
                "hsn_code": (hsn_from_form or item.get('hsn_code', '') or '').strip().upper(),
                "minimum_stock": 5,
                "location": "",
                "is_active": True,
                "created_at": now,
                "updated_at": now,
            }
            new_products.append(product_doc)

            if qty > 0:
                new_transactions.append({
                    "product_name": canonical,
                    "transaction_type": "INITIAL_STOCK",
                    "quantity": qty,
                    "previous_quantity": 0,
                    "new_quantity": qty,
                    "reason": "Initial product registration",
                    "performed_by": username,
                    "created_at": now,
                })
        else:
            existing_updates.append((product_name, qty, reason, username, now))

    db = get_db()

    if new_products:
        try:
            db.products.insert_many(new_products, ordered=False)
            created_count = len(new_products)
            success_count += created_count
            for p in new_products:
                invalidate_product_cache(p["product_name"])
                log_audit("PRODUCT_CREATE", username, p["product_name"], {"initial_stock": p["quantity"]})
        except Exception as e:
            errors.append(f"Bulk product creation failed: {str(e)}")

    if new_transactions:
        try:
            db.inventory_transactions.insert_many(new_transactions, ordered=False)
        except Exception:
            pass

    if existing_updates:
        product_names = [u[0] for u in existing_updates]
        products_cursor = db.products.find(
            {"product_name": {"$in": product_names}, "is_active": True},
            {"product_name": 1, "quantity": 1}
        )
        product_map = {p["product_name"]: p for p in products_cursor}

        bulk_tx = []
        for prod_name, qty, reason, user, ts in existing_updates:
            prod = product_map.get(prod_name)
            if not prod:
                errors.append(f"'{prod_name}' — Product not found or inactive")
                continue
            prev_qty = float(prod.get("quantity", 0))
            bulk_tx.append({
                "product_name": prod_name,
                "transaction_type": "STOCK_IN",
                "quantity": qty,
                "previous_quantity": prev_qty,
                "new_quantity": prev_qty + qty,
                "reason": (reason or "Inventory received").strip(),
                "performed_by": user,
                "created_at": ts,
            })
            success_count += 1

        if bulk_tx:
            try:
                db.inventory_transactions.insert_many(bulk_tx, ordered=False)

                for tx in bulk_tx:
                    db.products.update_one(
                        {"product_name": tx["product_name"]},
                        {"$inc": {"quantity": tx["quantity"]}, "$set": {"updated_at": now}}
                    )
                    invalidate_product_cache(tx["product_name"])
                    log_audit("STOCK_IN", tx["performed_by"], tx["product_name"],
                              {"qty_added": tx["quantity"], "new_qty": tx["new_quantity"]})
            except Exception as e:
                errors.append(f"Bulk stock-in failed: {str(e)}")

    if created_count > 0:
        flash(f"Created {created_count} new product(s).", "success")
    if success_count > 0:
        flash(f"Successfully stocked in {success_count} item(s).", "success")
    if errors:
        flash("Issues: " + "; ".join(errors), "warning")

    return redirect(url_for('inventory.view_transactions'))

@inventory_bp.route('/transactions')
@login_required
def view_transactions():
    p_name = request.args.get('product_name', '').strip()
    tx_type = request.args.get('type', '').strip()
    
    transactions = get_all_transactions(product_name=p_name, transaction_type=tx_type, limit=50)
    return render_template(
        'inventory/transactions.html',
        transactions=transactions,
        current_product=p_name,
        current_type=tx_type
    )

@inventory_bp.route('/transactions/export/excel')
@login_required
def export_transactions_excel():
    from datetime import datetime, timezone
    from flask import send_file
    from openpyxl.styles import Font, PatternFill
    from inventory_app.services.export_service import generate_excel_export, get_transaction_export_filters, build_export_subtitle
    
    filters = get_transaction_export_filters(request)
    
    transactions = get_all_transactions(product_name=filters['product_name'], transaction_type=filters['tx_type'], limit=1000)
    
    headers = [
        "Date & Time (UTC)", "Product Name", "Transaction Type", 
        "Qty Changed", "Previous Qty", "New Qty", 
        "Reason / Notes", "Performed By"
    ]
    
    data_rows = []
    for t in transactions:
        created_at_val = t.get('created_at', '')
        if hasattr(created_at_val, 'strftime'):
            created_at_val = created_at_val.strftime('%Y-%m-%d %H:%M:%S')
            
        data_rows.append([
            str(created_at_val),
            t.get('product_name', ''),
            str(t.get('transaction_type', '')),
            float(t.get('quantity', 0) or 0),
            float(t.get('previous_quantity', 0) or 0),
            float(t.get('new_quantity', 0) or 0),
            t.get('reason', ''),
            t.get('performed_by', '')
        ])

    font_family = "Segoe UI"
    status_style_map = {
        'STOCK_IN': (PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid"), Font(name=font_family, size=10, bold=True, color="166534")),
        'STOCK_OUT': (PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"), Font(name=font_family, size=10, bold=True, color="991B1B")),
        'ADJUSTMENT': (PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"), Font(name=font_family, size=10, bold=True, color="92400E"))
    }
    
    col_widths = {1: 22, 2: 24, 3: 18, 4: 15, 5: 15, 6: 15, 7: 28, 8: 18}
    number_formats = {4: '#,##0.00', 5: '#,##0.00', 6: '#,##0.00'}
    alignments = {1: 'center', 3: 'center', 4: 'right', 5: 'right', 6: 'right', 8: 'center'}

    subtitle = build_export_subtitle(f"Product Filter: {filters['product_name'] or 'All Products'} | Type Filter: {filters['tx_type'] or 'All Types'}", filters, len(transactions))

    buffer = generate_excel_export(
        sheet_title="Transaction Audit Log",
        banner_title="INVENTORY MANAGEMENT SYSTEM — TRANSACTION AUDIT REPORT",
        subtitle=subtitle,
        headers=headers,
        data_rows=data_rows,
        col_widths=col_widths,
        number_formats=number_formats,
        alignments=alignments,
        status_col_idx=3,
        status_style_map=status_style_map,
        summary_config={'label_col': 1, 'sum_cols': [4]}
    )
    
    return send_file(
        buffer,
        as_attachment=True,
        download_name="inventory_transactions_report.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@inventory_bp.route('/transactions/export/pdf')
@login_required
def export_transactions_pdf():
    from datetime import datetime, timezone
    from flask import send_file
    from reportlab.platypus import Paragraph
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from inventory_app.services.export_service import generate_pdf_export, get_transaction_export_filters, build_export_subtitle
    
    filters = get_transaction_export_filters(request)
    
    transactions = get_all_transactions(product_name=filters['product_name'], transaction_type=filters['tx_type'], limit=1000)
    
    styles = getSampleStyleSheet()
    cell_style = ParagraphStyle('TableCell', parent=styles['Normal'], fontName='Helvetica', fontSize=8.5)
    cell_center = ParagraphStyle('CellCenter', parent=cell_style, alignment=1)
    cell_right = ParagraphStyle('CellRight', parent=cell_style, alignment=2)
    cell_bold_right = ParagraphStyle('CellBoldRight', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=8.5, alignment=2)
    cell_bold_left = ParagraphStyle('CellBoldLeft', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=8.5, alignment=0)
    
    stock_in_pdf = sum(1 for t in transactions if t.get('transaction_type') == 'STOCK_IN')
    stock_out_pdf = sum(1 for t in transactions if t.get('transaction_type') == 'STOCK_OUT')
    tot_vol_pdf = sum(float(t.get('quantity', 0) or 0) for t in transactions)
    
    kpi_summary = [
        ["Total Audit Logs", "Stock In Events", "Stock Out Events", "Total Volume Changed"],
        [str(len(transactions)), str(stock_in_pdf), str(stock_out_pdf), f"{tot_vol_pdf:,.2f}"]
    ]
    
    headers = [
        "Date & Time (UTC)", "Product Name", "Type", 
        "Qty Changed", "Prev Qty", "New Qty", 
        "Reason / Notes", "Performed By"
    ]
    
    data_rows = []
    total_qty_sum = 0
    for t in transactions:
        created_at_val = t.get('created_at', '')
        if hasattr(created_at_val, 'strftime'):
            created_at_val = created_at_val.strftime('%Y-%m-%d %H:%M:%S')
            
        qty = float(t.get('quantity', 0) or 0)
        total_qty_sum += qty
        tx_type_val = str(t.get('transaction_type', ''))
        
        if tx_type_val == 'STOCK_IN':
            tx_html = f'<font color="#166534"><b>STOCK IN</b></font>'
        elif tx_type_val == 'STOCK_OUT':
            tx_html = f'<font color="#991B1B"><b>STOCK OUT</b></font>'
        elif tx_type_val == 'ADJUSTMENT':
            tx_html = f'<font color="#92400E"><b>ADJUSTMENT</b></font>'
        else:
            tx_html = f'<b>{tx_type_val}</b>'

        data_rows.append([
            Paragraph(str(created_at_val), cell_center),
            Paragraph(f"<b>{t.get('product_name', '')}</b>", cell_style),
            Paragraph(tx_html, cell_center),
            Paragraph(f"{qty:,.2f}", cell_bold_right),
            Paragraph(f"{float(t.get('previous_quantity', 0) or 0):,.2f}", cell_right),
            Paragraph(f"{float(t.get('new_quantity', 0) or 0):,.2f}", cell_bold_right),
            Paragraph(t.get('reason', ''), cell_style),
            Paragraph(t.get('performed_by', ''), cell_center)
        ])
        
    data_rows.append([
        Paragraph("TOTAL SUMMARY", cell_bold_left),
        Paragraph("", cell_style),
        Paragraph("", cell_style),
        Paragraph(f"{total_qty_sum:,.2f}", cell_bold_right),
        Paragraph("", cell_style),
        Paragraph("", cell_style),
        Paragraph("", cell_style),
        Paragraph("", cell_style)
    ])

    col_widths = [115, 140, 90, 75, 65, 65, 130, 90]
    subtitle = build_export_subtitle(f"Product Filter: {filters['product_name'] or 'All'} | Type Filter: {filters['tx_type'] or 'All'}", filters, len(transactions))

    buffer = generate_pdf_export(
        title="INVENTORY TRANSACTION AUDIT REPORT (A4 FORMATTED)",
        subtitle=subtitle,
        headers=headers,
        data_rows=data_rows,
        col_widths=col_widths,
        kpi_summary=kpi_summary
    )
    
    return send_file(
        buffer,
        as_attachment=True,
        download_name="inventory_transactions_report.pdf",
        mimetype="application/pdf"
    )
