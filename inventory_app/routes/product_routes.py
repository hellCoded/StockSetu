from flask import Blueprint, render_template, request, redirect, url_for, flash, session
from inventory_app.utils.decorators import login_required, roles_required, csrf_protected

product_bp = Blueprint('products', __name__)

@product_bp.route('/products')
@login_required
def list_products():
    from inventory_app.services.product_service import search_products, get_distinct_categories, get_distinct_locations
    from inventory_app.utils.pagination import Pagination
    query = request.args.get('q', '').strip()
    category = request.args.get('category', '').strip()
    location = request.args.get('location', '').strip()
    stock_status = request.args.get('status', '').strip()
    show_inactive = request.args.get('show_inactive', '0') == '1'
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 25, type=int)
    
    is_active_filter = None if show_inactive else True
    
    products, total_count = search_products(
        query=query,
        category=category,
        location=location,
        stock_status=stock_status,
        is_active=is_active_filter,
        page=page,
        per_page=per_page,
        return_total=True
    )
    
    pagination = Pagination(page=page, per_page=per_page, total=total_count)
    categories = get_distinct_categories()
    locations = get_distinct_locations()
    
    return render_template(
        'products/list.html',
        products=products,
        pagination=pagination,
        categories=categories,
        locations=locations,
        current_query=query,
        current_category=category,
        current_location=location,
        current_status=stock_status,
        show_inactive=show_inactive
    )

@product_bp.route('/products/export/excel')
@login_required
def export_excel():
    from datetime import datetime, timezone
    from flask import send_file
    from openpyxl.styles import Font, PatternFill
    from inventory_app.services.product_service import search_products
    from inventory_app.services.export_service import generate_excel_export, get_product_export_filters, build_export_subtitle
    
    filters = get_product_export_filters(request)
    
    products = search_products(
        query=filters['query'],
        category=filters['category'],
        location=filters['location'],
        stock_status=filters['stock_status'],
        is_active=filters['is_active']
    )
    
    headers = [
        "Product Name", "Category", "Stock Qty", "Unit", 
        "Unit Price (₹)", "Total Value (₹)", "Min Stock", 
        "Warehouse Location", "Stock Status", "Active Status"
    ]
    
    data_rows = []
    for p in products:
        qty = float(p.get('quantity', 0) or 0)
        price = float(p.get('price', 0) or 0)
        total_val = qty * price
        st_val = p.get('status', 'IN STOCK')
        data_rows.append([
            p.get('product_name', ''),
            p.get('category', ''),
            qty,
            p.get('unit', ''),
            price,
            total_val,
            5,
            p.get('location', '') or 'Unassigned',
            st_val,
            'Active' if p.get('is_active', True) else 'Inactive'
        ])

    font_family = "Segoe UI"
    status_style_map = {
        'IN STOCK': (PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid"), Font(name=font_family, size=10, bold=True, color="166534")),
        'LOW STOCK': (PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"), Font(name=font_family, size=10, bold=True, color="92400E")),
        'OUT OF STOCK': (PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"), Font(name=font_family, size=10, bold=True, color="991B1B"))
    }
    
    col_widths = {1: 26, 2: 18, 3: 14, 4: 12, 5: 16, 6: 18, 7: 14, 8: 20, 9: 16, 10: 14}
    number_formats = {3: '#,##0.00', 5: '"₹"#,##0.00', 6: '"₹"#,##0.00', 7: '#,##0.00'}
    alignments = {2: 'center', 3: 'right', 4: 'center', 5: 'right', 6: 'right', 7: 'right', 8: 'center', 9: 'center', 10: 'center'}

    subtitle = build_export_subtitle(f"Filter: {filters['query'] or 'All Products'}", filters, len(products))

    buffer = generate_excel_export(
        sheet_title="Inventory Catalog",
        banner_title="INVENTORY MANAGEMENT SYSTEM — PRODUCT CATALOG REPORT",
        subtitle=subtitle,
        headers=headers,
        data_rows=data_rows,
        col_widths=col_widths,
        number_formats=number_formats,
        alignments=alignments,
        status_col_idx=9,
        status_style_map=status_style_map,
        summary_config={'label_col': 1, 'sum_cols': [3, 6]}
    )
    
    return send_file(
        buffer,
        as_attachment=True,
        download_name="inventory_products_report.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@product_bp.route('/products/export/pdf')
@login_required
def export_pdf():
    from datetime import datetime, timezone
    from flask import send_file
    from reportlab.platypus import Paragraph
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from inventory_app.services.product_service import search_products
    from inventory_app.services.export_service import generate_pdf_export, get_product_export_filters, build_export_subtitle
    
    filters = get_product_export_filters(request)
    
    products = search_products(
        query=filters['query'],
        category=filters['category'],
        location=filters['location'],
        stock_status=filters['stock_status'],
        is_active=filters['is_active']
    )
    
    styles = getSampleStyleSheet()
    cell_style = ParagraphStyle('TableCell', parent=styles['Normal'], fontName='Helvetica', fontSize=8.5)
    cell_center = ParagraphStyle('CellCenter', parent=cell_style, alignment=1)
    cell_right = ParagraphStyle('CellRight', parent=cell_style, alignment=2)
    cell_bold_right = ParagraphStyle('CellBoldRight', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=8.5, alignment=2)
    cell_bold_left = ParagraphStyle('CellBoldLeft', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=8.5, alignment=0)
    
    total_units = sum(float(p.get('quantity', 0) or 0) for p in products)
    total_val = sum(float(p.get('quantity', 0) or 0) * float(p.get('price', 0) or 0) for p in products)
    low_pdf = sum(1 for p in products if p.get('status') == 'LOW STOCK')
    
    kpi_summary = [
        ["Total Products", "Total Units", "Total Catalog Value", "Low Stock Alerts"],
        [str(len(products)), f"{total_units:,.2f}", f"Rs. {total_val:,.2f}", str(low_pdf)]
    ]
    
    headers = [
        "Product Name", "Category", "Stock Qty", "Unit", 
        "Unit Price", "Total Value", "Min Stock", "Location", "Status"
    ]
    
    data_rows = []
    tot_qty_sum = 0
    tot_val_sum = 0
    for p in products:
        qty = float(p.get('quantity', 0) or 0)
        price = float(p.get('price', 0) or 0)
        item_val = qty * price
        tot_qty_sum += qty
        tot_val_sum += item_val
        st_val = p.get('status', 'IN STOCK')
        
        if st_val == 'IN STOCK':
            st_html = f'<font color="#166534"><b>{st_val}</b></font>'
        elif st_val == 'LOW STOCK':
            st_html = f'<font color="#92400E"><b>{st_val}</b></font>'
        elif st_val == 'OUT OF STOCK':
            st_html = f'<font color="#991B1B"><b>{st_val}</b></font>'
        else:
            st_html = st_val

        data_rows.append([
            Paragraph(f"<b>{p.get('product_name', '')}</b>", cell_style),
            Paragraph(p.get('category', ''), cell_center),
            Paragraph(f"{qty:,.2f}", cell_right),
            Paragraph(p.get('unit', ''), cell_center),
            Paragraph(f"Rs. {price:,.2f}", cell_right),
            Paragraph(f"Rs. {item_val:,.2f}", cell_bold_right),
            Paragraph(f"{5:,.2f}", cell_right),
            Paragraph(p.get('location', '') or 'Unassigned', cell_center),
            Paragraph(st_html, cell_center)
        ])
        
    data_rows.append([
        Paragraph("TOTAL SUMMARY", cell_bold_left),
        Paragraph("", cell_style),
        Paragraph(f"{tot_qty_sum:,.2f}", cell_bold_right),
        Paragraph("", cell_style),
        Paragraph("", cell_style),
        Paragraph(f"Rs. {tot_val_sum:,.2f}", cell_bold_right),
        Paragraph("", cell_style),
        Paragraph("", cell_style),
        Paragraph("", cell_style)
    ])

    col_widths = [160, 100, 60, 45, 80, 95, 60, 90, 80]
    subtitle = build_export_subtitle(f"Filter: {filters['query'] or 'All Catalog Products'}", filters, len(products))

    buffer = generate_pdf_export(
        title="INVENTORY CATALOG REPORT (A4 FORMATTED)",
        subtitle=subtitle,
        headers=headers,
        data_rows=data_rows,
        col_widths=col_widths,
        kpi_summary=kpi_summary
    )
    
    return send_file(
        buffer,
        as_attachment=True,
        download_name="inventory_products_report.pdf",
        mimetype="application/pdf"
    )

@product_bp.route('/products/add', methods=['GET', 'POST'])
@login_required
@roles_required('admin', 'inventory_manager')
@csrf_protected
def add_product():
    from inventory_app.services.product_service import create_product, get_distinct_categories
    if request.method == 'POST':
        product_data = {
            'product_name': request.form.get('product_name', ''),
            'category': request.form.get('category', ''),
            'description': request.form.get('description', ''),
            'quantity': request.form.get('quantity', 0),
            'unit': request.form.get('unit', ''),
            'price': request.form.get('price', 0),
            'gst_rate': request.form.get('gst_rate', 0),
            'hsn_code': request.form.get('hsn_code', ''),
            'location': request.form.get('location', '')
        }
        
        username = session.get('employee_id', 'System')
        success, msg, prod = create_product(product_data, performed_by=username)
        
        if success:
            flash(msg, "success")
            return redirect(url_for('products.view_product', product_name=prod['product_name']))
        else:
            flash(msg, "danger")
            categories = get_distinct_categories()
            return render_template('products/add.html', form_data=product_data, categories=categories)
            
    categories = get_distinct_categories()
    return render_template('products/add.html', categories=categories)

@product_bp.route('/products/<product_name>')
@login_required
def view_product(product_name):
    from inventory_app.services.product_service import get_product_by_name
    from inventory_app.services.inventory_service import get_product_transactions
    product = get_product_by_name(product_name)
    if not product:
        flash(f"Product '{product_name}' not found.", "warning")
        return redirect(url_for('products.list_products'))
        
    transactions = get_product_transactions(product['product_name'])
    return render_template('products/detail.html', product=product, transactions=transactions)

@product_bp.route('/products/<product_name>/edit', methods=['GET', 'POST'])
@login_required
@roles_required('admin', 'inventory_manager')
@csrf_protected
def edit_product(product_name):
    from inventory_app.services.product_service import get_product_by_name, update_product, get_distinct_categories
    product = get_product_by_name(product_name)
    if not product:
        flash(f"Product '{product_name}' not found.", "warning")
        return redirect(url_for('products.list_products'))
        
    if request.method == 'POST':
        update_data = {
            'category': request.form.get('category', ''),
            'description': request.form.get('description', ''),
            'unit': request.form.get('unit', ''),
            'price': request.form.get('price', 0),
            'gst_rate': request.form.get('gst_rate', 0),
            'hsn_code': request.form.get('hsn_code', ''),
            'location': request.form.get('location', '')
        }
        
        username = session.get('employee_id', 'System')
        success, msg, updated_prod = update_product(product['product_name'], update_data, performed_by=username)
        
        if success:
            flash(msg, "success")
            return redirect(url_for('products.view_product', product_name=updated_prod['product_name']))
        else:
            flash(msg, "danger")
            categories = get_distinct_categories()
            return render_template('products/edit.html', product=product, categories=categories)
            
    categories = get_distinct_categories()
    return render_template('products/edit.html', product=product, categories=categories)

@product_bp.route('/products/<product_name>/rename', methods=['GET', 'POST'])
@login_required
@roles_required('admin')
@csrf_protected
def rename_product(product_name):
    from inventory_app.services.product_service import get_product_by_name, rename_product as rename_product_svc
    product = get_product_by_name(product_name)
    if not product:
        flash(f"Product '{product_name}' not found.", "warning")
        return redirect(url_for('products.list_products'))
        
    if request.method == 'POST':
        new_name = request.form.get('new_product_name', '')
        username = session.get('employee_id', 'System')
        
        success, msg = rename_product_svc(product['product_name'], new_name, performed_by=username)
        if success:
            flash(msg, "success")
            return redirect(url_for('products.view_product', product_name=new_name))
        else:
            flash(msg, "danger")
            return render_template('products/rename.html', product=product, new_product_name=new_name)
            
    return render_template('products/rename.html', product=product)

@product_bp.route('/products/<product_name>/toggle-active', methods=['POST'])
@login_required
@roles_required('admin', 'inventory_manager')
@csrf_protected
def toggle_active(product_name):
    from inventory_app.services.product_service import toggle_product_active
    username = session.get('username', 'System')
    success, msg, _ = toggle_product_active(product_name, performed_by=username)
    if success:
        flash(msg, "success")
    else:
        flash(msg, "danger")
    return redirect(url_for('products.view_product', product_name=product_name))


@product_bp.route('/api/global-search')
@login_required
def api_global_search():
    import re
    from flask import jsonify
    from inventory_app.database import get_db
    
    q = request.args.get('q', '').strip()
    if not q:
        return jsonify({"products": [], "bills": []})
        
    db = get_db()
    regex = re.compile(re.escape(q), re.IGNORECASE)
    
    prods = list(db.products.find(
        {"$or": [
            {"product_name": {"$regex": regex}},
            {"category": {"$regex": regex}},
            {"hsn_code": {"$regex": regex}}
        ]}
    ).limit(6))
    
    prod_results = [{
        "name": p.get("product_name"),
        "category": p.get("category", ""),
        "quantity": p.get("quantity", 0),
        "unit": p.get("unit", "pcs"),
        "price": p.get("price", 0),
        "hsn": p.get("hsn_code", ""),
        "status": "In Stock" if (p.get("quantity", 0) > (p.get("minimum_stock", 5))) else ("Low Stock" if p.get("quantity", 0) > 0 else "Out of Stock"),
        "url": url_for('products.view_product', product_name=p.get("product_name"))
    } for p in prods]
    
    bills = list(db.invoices.find(
        {"$or": [
            {"bill_number": {"$regex": regex}},
            {"customer_name": {"$regex": regex}},
            {"customer_phone": {"$regex": regex}}
        ]}
    ).sort("created_at", -1).limit(5))
    
    bill_results = [{
        "id": str(b.get("_id")),
        "number": b.get("bill_number"),
        "customer": b.get("customer_name"),
        "grand_total": b.get("grand_total", 0),
        "status": b.get("payment_status", "PAID"),
        "url": url_for('billing.view_bill', bill_id=str(b.get("_id")))
    } for b in bills]
    
    return jsonify({
        "products": prod_results,
        "bills": bill_results
    })

