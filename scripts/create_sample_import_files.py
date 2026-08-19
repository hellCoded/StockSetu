import os
import csv
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

SAMPLE_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'sample_import_files')
os.makedirs(SAMPLE_DIR, exist_ok=True)

# Common styles
thin_border = Border(
    left=Side(style='thin', color='CBD5E1'),
    right=Side(style='thin', color='CBD5E1'),
    top=Side(style='thin', color='CBD5E1'),
    bottom=Side(style='thin', color='CBD5E1')
)

def create_excel_and_csv(base_filename, sheet_title, headers, rows, header_color="1E3A8A"):
    # 1. CSV
    csv_path = os.path.join(SAMPLE_DIR, f"{base_filename}.csv")
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(rows)

    # 2. XLSX
    xlsx_path = os.path.join(SAMPLE_DIR, f"{base_filename}.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]
    ws.append(headers)

    h_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
    h_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.fill = h_fill
        cell.font = h_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in rows:
        ws.append(row)

    for row in ws.iter_rows(min_row=2, max_row=len(rows)+1, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    wb.save(xlsx_path)


def create_supplier_pdf(filename, supplier_name, invoice_no, invoice_date, headers, rows, brand_color=colors.HexColor("#1e3a8a")):
    pdf_path = os.path.join(SAMPLE_DIR, f"{filename}.pdf")
    doc = SimpleDocTemplate(
        pdf_path,
        pagesize=A4,
        rightMargin=36,
        leftMargin=36,
        topMargin=36,
        bottomMargin=36
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'SupplierTitle',
        parent=styles['Heading1'],
        fontSize=18,
        leading=22,
        textColor=brand_color,
        fontName='Helvetica-Bold'
    )
    meta_style = ParagraphStyle(
        'MetaText',
        parent=styles['Normal'],
        fontSize=9,
        leading=13,
        textColor=colors.HexColor("#334155")
    )

    story = []

    # Header info table
    header_data = [
        [
            Paragraph(f"<b>{supplier_name}</b><br/><font size=8 color='#64748b'>Authorized Wholesale Supplier & Distributor</font>", title_style),
            Paragraph(f"<b>TAX INVOICE</b><br/>Invoice #: <b>{invoice_no}</b><br/>Date: {invoice_date}<br/>GSTIN: 27AAACR1234F1Z5", meta_style)
        ]
    ]
    t_header = Table(header_data, colWidths=[320, 200])
    t_header.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('ALIGN', (1,0), (1,0), 'RIGHT'),
    ]))
    story.append(t_header)
    story.append(Spacer(1, 14))

    # Billed To
    billed_data = [
        [
            Paragraph("<b>Billed To:</b><br/>StockSetu Central Warehouse<br/>Industrial Estate, Sector 4<br/>GSTIN: 27AABCS5678K1Z2", meta_style),
            Paragraph("<b>Place of Supply:</b> Maharashtra (27)<br/><b>Payment Terms:</b> Net 30 Days<br/><b>Vehicle No:</b> MH-12-Q-4589", meta_style)
        ]
    ]
    t_billed = Table(billed_data, colWidths=[320, 200])
    t_billed.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('BACKGROUND', (0,0), (-1,-1), colors.HexColor("#f8fafc")),
        ('BOX', (0,0), (-1,-1), 0.5, colors.HexColor("#cbd5e1")),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('LEFTPADDING', (0,0), (-1,-1), 8),
        ('RIGHTPADDING', (0,0), (-1,-1), 8),
    ]))
    story.append(t_billed)
    story.append(Spacer(1, 16))

    # Product Table
    table_data = [[h for h in headers]]
    for r in rows:
        table_data.append([str(c) for c in r])

    # Table layout: give more width to item name and use landscape or balanced widths
    num_cols = len(headers)
    col_widths = [185, 45, 45, 55, 45, 55, 95] if num_cols == 7 else [160, 40, 40, 50, 45, 45, 70, 75]

    t_items = Table(table_data, colWidths=col_widths[:num_cols], repeatRows=1)
    t_items.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), brand_color),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 9),
        ('ALIGN', (0,0), (-1,0), 'CENTER'),
        ('ALIGN', (0,1), (0,-1), 'LEFT'),
        ('ALIGN', (1,1), (-1,-1), 'CENTER'),
        ('FONTNAME', (0,1), (-1,-1), 'Helvetica'),
        ('FONTSIZE', (0,1), (-1,-1), 8.5),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
        ('TOPPADDING', (0,0), (-1,-1), 5),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#cbd5e1")),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor("#f8fafc")]),
    ]))
    story.append(t_items)
    story.append(Spacer(1, 18))

    # Summary Footer
    footer_text = Paragraph("<font size=8 color='#64748b'>This is a computer generated invoice for StockSetu Bulk Stock-In demonstrations and live system verification.</font>", meta_style)
    story.append(footer_text)

    doc.build(story)


# ==============================================================================
# DATASET DEFINITIONS
# ==============================================================================

# SET 1: Electrical & Lighting Supplies
headers_prod = ["Product Name", "Quantity", "Unit", "Price", "GST Rate", "HSN Code", "Category"]

rows_elec = [
    ["Havells 2.5 sq mm Copper Wire 90m", 40, "rolls", 1450.00, 18, "8544", "Electrical"],
    ["Anchor Roma 6A 1-Way Modular Switch", 200, "pcs", 38.00, 18, "8536", "Electrical"],
    ["Schneider 32A Double Pole MCB", 30, "pcs", 420.00, 18, "8536", "Electrical"],
    ["Philips 12W Cool Daylight LED Bulb", 120, "pcs", 95.00, 18, "9405", "Electrical"],
    ["Polycab 4-Core 1.5 sq mm Cable 100m", 15, "rolls", 2850.00, 18, "8544", "Electrical"],
    ["Crompton 1200mm High Speed Ceiling Fan", 12, "pcs", 1750.00, 18, "8414", "Electrical"],
]
create_excel_and_csv("supplier_invoice_electrical", "Electrical Supplies", headers_prod, rows_elec, "1E3A8A")
create_supplier_pdf("supplier_invoice_electrical", "Havells & Anchor Electrical Distributors", "INV-ELEC-2026-088", "19-Aug-2026", headers_prod, rows_elec, colors.HexColor("#1e3a8a"))


# SET 2: Plumbing, Sanitary & Hardware Supplies
rows_plumb = [
    ["Supreme 1 Inch PVC Pipe (3m)", 100, "pcs", 210.00, 18, "3917", "Plumbing"],
    ["Astral CPVC 3/4 Inch Ball Valve", 50, "pcs", 185.00, 18, "8481", "Plumbing"],
    ["Jaquar Continental Brass Bib Cock Tap", 25, "pcs", 850.00, 18, "8481", "Plumbing"],
    ["Sintex 1000L Triple Layer Water Tank", 6, "pcs", 4500.00, 18, "3925", "Plumbing"],
    ["Stanley 5-Meter Steel Measuring Tape", 20, "pcs", 190.00, 18, "9017", "Tools & Equipment"],
    ["Godrej 6-Lever Brass Padlock 65mm", 35, "pcs", 320.00, 18, "8301", "Hardware & Fasteners"],
]
create_excel_and_csv("supplier_invoice_plumbing_hardware", "Plumbing & Hardware", headers_prod, rows_plumb, "0F766E")
create_supplier_pdf("supplier_invoice_plumbing_hardware", "Apex Plumbing & Hardware Wholesalers", "INV-PLUMB-2026-104", "19-Aug-2026", headers_prod, rows_plumb, colors.HexColor("#0f766e"))


# SET 3: Heavy Building Materials, Cement & Paints
rows_bldg = [
    ["UltraTech Cement 53 Grade (50kg)", 150, "bags", 385.00, 28, "2523", "Cement"],
    ["Tata Tiscon 16mm Fe500D TMT Bar", 60, "pcs", 890.00, 18, "7214", "Building Materials"],
    ["Asian Paints Apex Ultima White 20L", 18, "liters", 4350.00, 18, "3209", "Paints & Chemicals"],
    ["Dr Fixit Pidiproof LW+ Waterproof 5L", 30, "liters", 680.00, 18, "3824", "Paints & Chemicals"],
    ["Bosch 600W Professional Angle Grinder", 8, "pcs", 2450.00, 18, "8467", "Tools & Equipment"],
    ["Karam Full Body Safety Harness Belt", 15, "pcs", 1150.00, 18, "6307", "Safety Gear"],
]
create_excel_and_csv("supplier_invoice_building_materials", "Building Materials", headers_prod, rows_bldg, "C2410C")
create_supplier_pdf("supplier_invoice_building_materials", "Bharat Cement & Industrial Paint Corp", "INV-BLDG-2026-512", "19-Aug-2026", headers_prod, rows_bldg, colors.HexColor("#c2410c"))


# SET 4: Staff Bulk Import Sets (Branch A & Branch B)
staff_headers = ["Employee ID", "Full Name", "Phone No", "Email"]

staff_branch_a = [
    ["EMP-2001", "Karan Singhania", "9811122233", "karan.s@stocksetu.com"],
    ["EMP-2002", "Neha Deshmukh", "9811122234", "neha.d@stocksetu.com"],
    ["EMP-2003", "Rohan Mehta", "9811122235", "rohan.m@stocksetu.com"],
    ["EMP-2004", "Ananya Roy", "9811122236", "ananya.r@stocksetu.com"],
    ["EMP-2005", "Deepak Joshi", "9811122237", "deepak.j@stocksetu.com"],
]
create_excel_and_csv("staff_bulk_import_branch_a", "Branch A Staff", staff_headers, staff_branch_a, "4338CA")

staff_branch_b = [
    ["EMP-3001", "Sanjay Nambiar", "9822233344", "sanjay.n@stocksetu.com"],
    ["EMP-3002", "Pooja Hegde", "9822233345", "pooja.h@stocksetu.com"],
    ["EMP-3003", "Manish Kulkarni", "9822233346", "manish.k@stocksetu.com"],
    ["EMP-3004", "Divya Pillai", "9822233347", "divya.p@stocksetu.com"],
    ["EMP-3005", "Aditya Kapoor", "9822233348", "aditya.k@stocksetu.com"],
]
create_excel_and_csv("staff_bulk_import_branch_b", "Branch B Staff", staff_headers, staff_branch_b, "7C3AED")

print("All sample files (PDF, XLSX, CSV) generated successfully in:", SAMPLE_DIR)
